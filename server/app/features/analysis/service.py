from __future__ import annotations

import base64
import csv
import html
import json
import math
import mimetypes
import re
import subprocess
import threading
import time
import uuid
import zipfile
from datetime import date, datetime, time as dt_time
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook
import requests

from app.core.config import settings
from app.core.errors import AppError
from app.features.focus.repository import repository as focus_repository
from app.features.params.service import service as params_service

from .hybrid_adapter import hybrid_adapter
from .repository import repository
from .schemas import (
    AnalysisAsset,
    AnalysisJobCreateResponse,
    AnalysisJobDetailResponse,
    AnalysisJobListItem,
    AnalysisJobListResponse,
    AnalysisParserSummary,
    AnalysisReportResponse,
    AnalysisResult,
    AnalysisRiskNode,
)


SPREADSHEET_EXTENSIONS = {".xlsx", ".csv"}
TEXT_EXTENSIONS = {".txt", ".md", ".json"}
HTML_EXTENSIONS = {".html", ".htm"}
HTTP_RETRY_DELAYS = (0.0, 1.5, 4.0)
REPORT_TITLE = "金融欺诈检测分析报告"
STANDARD_TRANSACTION_FIELDS = (
    "jylsxh",
    "zhdh",
    "dfzh",
    "jdbj",
    "jyje",
    "zhye",
    "dfhh",
    "jyrq",
    "jysj",
    "jyqd",
    "dfmccd",
    "xb",
    "年龄",
)
SPREADSHEET_FIELD_KEYWORDS = {
    "jylsxh": ("交易流水序号", "流水号", "流水", "序号", "jylsxh"),
    "zhdh": ("账户代号", "付款账户", "付款账号", "账户", "账号", "开户账号", "zhdh"),
    "dfzh": ("对方账户", "收款账户", "收款账号", "对手账户", "对方账号", "dfzh"),
    "jdbj": ("借贷标记", "借贷", "收付标志", "交易方向", "jdbj"),
    "jyje": ("交易金额", "金额", "发生额", "交易额", "jyje"),
    "zhye": ("账户余额", "余额", "可用余额", "zhye"),
    "dfhh": ("对方行号", "对方银行", "银行行号", "联行号", "dfhh"),
    "jyrq": ("交易日期", "日期", "入账日期", "jyrq"),
    "jysj": ("交易时间", "时间", "时刻", "jysj"),
    "jyqd": ("交易渠道", "渠道", "终端渠道", "jyqd"),
    "dfmccd": ("对方陌生程度", "陌生程度", "dfmccd"),
    "xb": ("性别", "xb"),
    "年龄": ("年龄", "age"),
}


class AnalysisService:
    def __init__(self) -> None:
        self.storage_dir = Path(settings.analysis_storage_dir)
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        self.http = requests.Session()
        self.http.trust_env = settings.analysis_http_trust_env

    def create_job(
        self,
        *,
        current_user: dict,
        text_payload: str,
        files: list[UploadFile],
    ) -> AnalysisJobCreateResponse:
        if not text_payload.strip() and not files:
            raise AppError("请先上传图片、文本或文件", status_code=400, code="EMPTY_ANALYSIS_INPUT")

        job_id = self._generate_job_id()
        job_dir = self.storage_dir / job_id
        job_dir.mkdir(parents=True, exist_ok=True)

        repository.create_job(job_id=job_id, created_by=current_user["username"], status="pending")

        if text_payload.strip():
            text_path = job_dir / "text_payload.txt"
            text_path.write_text(text_payload.strip(), encoding="utf-8")
            repository.add_asset(
                asset_id=self._generate_asset_id(),
                job_id=job_id,
                asset_type="text",
                original_name="text_payload.txt",
                mime_type="text/plain",
                local_path=str(text_path),
                size_bytes=text_path.stat().st_size,
            )

        for upload in files:
            asset_id = self._generate_asset_id()
            target_path = job_dir / f"{asset_id}_{Path(upload.filename or 'upload.bin').name}"
            content = upload.file.read()
            target_path.write_bytes(content)
            repository.add_asset(
                asset_id=asset_id,
                job_id=job_id,
                asset_type=self._classify_asset_type(target_path),
                original_name=upload.filename or target_path.name,
                mime_type=upload.content_type or self._guess_mime_type(target_path),
                local_path=str(target_path),
                size_bytes=len(content),
            )

        worker = threading.Thread(target=self._process_job, args=(job_id,), daemon=True)
        worker.start()
        return AnalysisJobCreateResponse(job_id=job_id, status="pending")

    def get_job(self, job_id: str, current_user: dict) -> AnalysisJobDetailResponse:
        job = repository.get_job(job_id)
        if not job:
            raise AppError("分析任务不存在", status_code=404, code="ANALYSIS_JOB_NOT_FOUND")

        self._ensure_job_access(job, current_user)
        assets = [self._to_asset_schema(item) for item in repository.list_assets(job_id)]
        parser_summary = (
            AnalysisParserSummary(**job["parser_summary_json"]) if job.get("parser_summary_json") else None
        )
        result_payload = self._normalize_result_payload(job.get("result_json"))
        result = AnalysisResult(**result_payload) if result_payload else None
        return AnalysisJobDetailResponse(
            job_id=job["job_id"],
            status=job["status"],
            created_by=job["created_by"],
            created_at=job["created_at"],
            started_at=job["started_at"],
            finished_at=job["finished_at"],
            error_message=job["error_message"],
            assets=assets,
            parser_summary=parser_summary,
            result=result,
            report_ready=bool(job.get("report_path")),
        )

    def list_jobs(self, current_user: dict) -> AnalysisJobListResponse:
        jobs = repository.list_jobs(created_by=None if current_user["role"] == "admin" else current_user["username"])
        return AnalysisJobListResponse(
            items=[
                AnalysisJobListItem(
                    job_id=item["job_id"],
                    status=item["status"],
                    created_by=item["created_by"],
                    created_at=item["created_at"],
                    finished_at=item["finished_at"],
                    risk_level=(item["result_json"] or {}).get("risk_level"),
                    confidence=(item["result_json"] or {}).get("confidence"),
                )
                for item in jobs
            ]
        )

    def get_report(self, job_id: str, current_user: dict) -> AnalysisReportResponse:
        job = repository.get_job(job_id)
        if not job:
            raise AppError("分析任务不存在", status_code=404, code="ANALYSIS_JOB_NOT_FOUND")
        self._ensure_job_access(job, current_user)
        report_path = self._ensure_report_files(job_id, job)
        return AnalysisReportResponse(
            job_id=job_id,
            title=job.get("report_title") or REPORT_TITLE,
            html=Path(report_path).read_text(encoding="utf-8"),
        )

    def get_report_pdf_path(self, job_id: str, current_user: dict) -> Path:
        job = repository.get_job(job_id)
        if not job:
            raise AppError("分析任务不存在", status_code=404, code="ANALYSIS_JOB_NOT_FOUND")
        self._ensure_job_access(job, current_user)
        report_path = self._ensure_report_files(job_id, job)
        pdf_path = Path(report_path).with_suffix(".pdf")
        if not pdf_path.exists():
            raise AppError("PDF 报告尚未生成", status_code=404, code="ANALYSIS_PDF_NOT_FOUND")
        return pdf_path

    def _process_job(self, job_id: str) -> None:
        repository.update_job(job_id, {"status": "processing", "started_at": self._now_iso()})
        try:
            assets = repository.list_assets(job_id)
            parser_summary = AnalysisParserSummary()
            bundle = {
                "job_id": job_id,
                "documents_markdown": [],
                "documents_json": [],
                "structured_tables": [],
                "plain_texts": [],
                "source_meta": [],
            }

            mineru_assets = [asset for asset in assets if asset["asset_type"] == "document"]
            if mineru_assets:
                mineru_results = self._parse_with_mineru(mineru_assets, parser_summary)
                bundle["documents_markdown"].extend(mineru_results["documents_markdown"])
                bundle["documents_json"].extend(mineru_results["documents_json"])
                bundle["source_meta"].extend(mineru_results["source_meta"])

            for asset in assets:
                asset_path = Path(asset["local_path"])
                if asset["asset_type"] == "spreadsheet":
                    spreadsheet_payload = self._parse_spreadsheet(asset_path)
                    parser_summary.spreadsheet_assets += 1
                    bundle["structured_tables"].extend(spreadsheet_payload["tables"])
                    bundle["source_meta"].append(
                        {
                            "asset_id": asset["asset_id"],
                            "asset_type": asset["asset_type"],
                            "original_name": asset["original_name"],
                        }
                    )
                    continue

                if asset["asset_type"] == "text":
                    parser_summary.plain_text_assets += 1
                    bundle["plain_texts"].append(
                        {
                            "asset_id": asset["asset_id"],
                            "content": asset_path.read_text(encoding="utf-8", errors="ignore"),
                        }
                    )
                    bundle["source_meta"].append(
                        {
                            "asset_id": asset["asset_id"],
                            "asset_type": asset["asset_type"],
                            "original_name": asset["original_name"],
                        }
                    )

            normalized = self._normalize_bundle(bundle, parser_summary)
            normalized = self._inject_focus_history(normalized, job_id)
            result = self._analyze_bundle(normalized, parser_summary, self.storage_dir / job_id)
            result = self._apply_focus_targets(normalized, result)
            report_path = self._render_report(job_id=job_id, assets=assets, normalized=normalized, result=result)

            repository.update_job(
                job_id,
                {
                    "status": "completed",
                    "parser_summary_json": parser_summary.model_dump(),
                    "normalized_json": normalized,
                    "result_json": result,
                    "report_path": str(report_path),
                    "report_title": REPORT_TITLE,
                    "finished_at": self._now_iso(),
                    "error_message": None,
                },
            )
        except Exception as exc:
            repository.update_job(
                job_id,
                {
                    "status": "failed",
                    "error_message": str(exc),
                    "finished_at": self._now_iso(),
                },
            )

    def _parse_with_mineru(self, assets: list[dict], parser_summary: AnalysisParserSummary) -> dict:
        if not settings.mineru_api_token:
            parser_summary.warnings.append("未配置 MinerU token，文档解析已退回本地摘要模式。")
            return self._fallback_parse_documents(assets)

        try:
            html_assets = [asset for asset in assets if Path(asset["local_path"]).suffix.lower() in HTML_EXTENSIONS]
            other_assets = [asset for asset in assets if asset not in html_assets]
            merged = {
                "documents_markdown": [],
                "documents_json": [],
                "source_meta": [],
            }

            if html_assets:
                html_results = self._submit_mineru_batch(html_assets, model_version="MinerU-HTML")
                parser_summary.mineru_documents += len(html_results)
                html_payload = self._collect_mineru_results(html_results)
                merged["documents_markdown"].extend(html_payload["documents_markdown"])
                merged["documents_json"].extend(html_payload["documents_json"])
                merged["source_meta"].extend(html_payload["source_meta"])

            if other_assets:
                document_results = self._submit_mineru_batch(other_assets, model_version="vlm")
                parser_summary.mineru_documents += len(document_results)
                document_payload = self._collect_mineru_results(document_results)
                merged["documents_markdown"].extend(document_payload["documents_markdown"])
                merged["documents_json"].extend(document_payload["documents_json"])
                merged["source_meta"].extend(document_payload["source_meta"])

            return merged
        except Exception as exc:
            parser_summary.warnings.append(f"MinerU 解析失败，已退回本地摘要模式：{exc}")
            return self._fallback_parse_documents(assets)

    def _submit_mineru_batch(self, assets: list[dict], *, model_version: str) -> list[dict]:
        body = {
            "files": [{"name": asset["original_name"], "data_id": asset["asset_id"]} for asset in assets],
            "model_version": model_version,
            "enable_formula": True,
            "enable_table": True,
            "language": "ch",
            "extra_formats": ["html"],
        }
        response = self._request_json(
            method="POST",
            url=f"{settings.mineru_api_base}/file-urls/batch",
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {settings.mineru_api_token}",
            },
            data=body,
            timeout=settings.analysis_llm_timeout_seconds,
        )
        if response.get("code") != 0:
            raise RuntimeError(response.get("msg") or "MinerU 上传链接申请失败")

        for asset, upload_url in zip(assets, response["data"]["file_urls"], strict=True):
            self._put_file(upload_url, Path(asset["local_path"]))

        return self._poll_mineru_batch(response["data"]["batch_id"])

    def _poll_mineru_batch(self, batch_id: str) -> list[dict]:
        deadline = time.time() + settings.mineru_poll_timeout_seconds
        last_payload: list[dict] = []
        while time.time() < deadline:
            response = self._request_json(
                method="GET",
                url=f"{settings.mineru_api_base}/extract-results/batch/{batch_id}",
                headers={
                    "Authorization": f"Bearer {settings.mineru_api_token}",
                    "Accept": "*/*",
                },
                timeout=settings.analysis_llm_timeout_seconds,
            )
            if response.get("code") != 0:
                raise RuntimeError(response.get("msg") or "MinerU 查询失败")

            last_payload = response.get("data", {}).get("extract_result", []) or []
            states = {item.get("state") for item in last_payload}
            if states and states.issubset({"done", "failed"}):
                return last_payload
            time.sleep(settings.mineru_poll_interval_seconds)

        raise TimeoutError(f"MinerU 轮询超时，最后状态：{last_payload}")

    def _collect_mineru_results(self, extract_results: list[dict]) -> dict:
        documents_markdown: list[dict] = []
        documents_json: list[dict] = []
        source_meta: list[dict] = []

        for item in extract_results:
            data_id = item.get("data_id")
            file_name = item.get("file_name")
            if item.get("state") != "done" or not item.get("full_zip_url"):
                documents_markdown.append(
                    {
                        "asset_id": data_id,
                        "source_type": "document",
                        "markdown": f"# {file_name}\n\nMinerU 未能完成解析：{item.get('err_msg') or item.get('state')}",
                    }
                )
                source_meta.append({"asset_id": data_id, "asset_type": "document", "original_name": file_name})
                continue

            archive_path = self._write_temp_zip(data_id or uuid.uuid4().hex, self._download_bytes(item["full_zip_url"], timeout=settings.analysis_llm_timeout_seconds))
            markdown_text = ""
            structured_payload: dict = {}
            with zipfile.ZipFile(archive_path) as archive:
                for name in archive.namelist():
                    lowered = name.lower()
                    if lowered.endswith("full.md") or lowered.endswith(".md"):
                        markdown_text = archive.read(name).decode("utf-8", errors="ignore")
                    elif lowered.endswith("content_list.json") or lowered.endswith("model.json") or lowered.endswith(".json"):
                        try:
                            structured_payload[name] = json.loads(archive.read(name).decode("utf-8", errors="ignore"))
                        except json.JSONDecodeError:
                            continue

            documents_markdown.append(
                {
                    "asset_id": data_id,
                    "source_type": "document",
                    "markdown": markdown_text or f"# {file_name}\n\nMinerU 已完成解析，但未返回 markdown 正文。",
                }
            )
            documents_json.append(
                {
                    "asset_id": data_id,
                    "source_type": "document",
                    "payload": structured_payload,
                }
            )
            source_meta.append({"asset_id": data_id, "asset_type": "document", "original_name": file_name})

        return {
            "documents_markdown": documents_markdown,
            "documents_json": documents_json,
            "source_meta": source_meta,
        }

    def _fallback_parse_documents(self, assets: list[dict]) -> dict:
        documents_markdown: list[dict] = []
        documents_json: list[dict] = []
        source_meta: list[dict] = []
        for asset in assets:
            asset_path = Path(asset["local_path"])
            documents_markdown.append(
                {
                    "asset_id": asset["asset_id"],
                    "source_type": "document",
                    "markdown": f"# {asset['original_name']}\n\n本地已接收文件，等待云端文档解析。\n\n- 文件大小：{asset['size_bytes']} 字节\n- 文件类型：{asset['mime_type']}",
                }
            )
            documents_json.append(
                {
                    "asset_id": asset["asset_id"],
                    "source_type": "document",
                    "payload": {
                        "file_name": asset["original_name"],
                        "extension": asset_path.suffix.lower(),
                        "size_bytes": asset["size_bytes"],
                    },
                }
            )
            source_meta.append(
                {
                    "asset_id": asset["asset_id"],
                    "asset_type": "document",
                    "original_name": asset["original_name"],
                }
            )
        return {
            "documents_markdown": documents_markdown,
            "documents_json": documents_json,
            "source_meta": source_meta,
        }

    def _parse_spreadsheet(self, path: Path) -> dict:
        if path.suffix.lower() == ".csv":
            return {"tables": [self._parse_csv_table(path)]}
        return {"tables": self._parse_xlsx_tables(path)}

    def _parse_csv_table(self, path: Path) -> dict:
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = list(csv.reader(handle))
        headers = reader[0] if reader else []
        rows = reader[1:26] if len(reader) > 1 else []
        return {
            "asset_id": path.stem,
            "sheet_name": "Sheet1",
            "columns": headers,
            "rows": [dict(zip(headers, row, strict=False)) for row in rows],
            "markdown_summary": self._table_to_markdown("Sheet1", headers, rows[:10]),
        }

    def _parse_xlsx_tables(self, path: Path) -> list[dict]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        tables: list[dict] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(cell or "").strip() for cell in rows[0]]
            data_rows = rows[1:]
            structured_rows = []
            for row in data_rows:
                structured_rows.append(
                    {
                        headers[index] if index < len(headers) else f"column_{index + 1}": row[index]
                        for index in range(len(row))
                    }
                )
            tables.append(
                {
                    "asset_id": path.stem,
                    "sheet_name": sheet.title,
                    "columns": headers,
                    "rows": structured_rows,
                    "markdown_summary": self._table_to_markdown(sheet.title, headers, data_rows[:10]),
                }
            )
        workbook.close()
        return tables

    def _normalize_spreadsheets(self, tables: list[dict], parser_summary: AnalysisParserSummary) -> dict:
        transactions: list[dict] = []
        candidates: list[dict] = []
        accounts: set[str] = set()
        timestamps: set[str] = set()
        amounts: list[float] = []
        mappings: list[dict] = []
        risk_signals: list[str] = []

        for table in tables:
            mapping = self._infer_spreadsheet_mapping(table, parser_summary)
            mappings.append(
                {
                    "asset_id": table.get("asset_id"),
                    "sheet_name": table.get("sheet_name"),
                    "column_mapping": mapping.get("column_mapping", {}),
                    "static_mapping": mapping.get("static_mapping", {}),
                }
            )
            standardized_rows = self._apply_spreadsheet_mapping(table, mapping)
            for row in standardized_rows:
                transactions.append(row)
                account = self._stringify(row.get("zhdh")).strip()
                counterpart = self._stringify(row.get("dfzh")).strip()
                if account:
                    accounts.add(account)
                if counterpart:
                    accounts.add(counterpart)
                amount = self._safe_float(row.get("jyje"))
                if amount:
                    amounts.append(amount)
                timestamp = " ".join(
                    part
                    for part in [self._stringify(row.get("jyrq")).strip(), self._stringify(row.get("jysj")).strip()]
                    if part
                ).strip()
                if timestamp:
                    timestamps.add(timestamp)
                candidates.append(
                    {
                        "payer_account": account,
                        "receiver_account": counterpart,
                        "amount": amount,
                        "timestamp": timestamp,
                        "channel": self._stringify(row.get("jyqd")) or "Excel结构化表格",
                        "description": f"{table['sheet_name']} 表单候选交易",
                        "evidence_refs": [table["sheet_name"]],
                    }
                )

        if any(amount >= 50000 for amount in amounts):
            risk_signals.append("存在大额转账记录")
        if len(accounts) >= 3:
            risk_signals.append("检测到多账户关联")
        if transactions:
            risk_signals.append("已完成 Excel 表头映射与全量结构化")

        return {
            "transaction_candidates": candidates,
            "standardized_transactions": transactions,
            "spreadsheet_mappings": mappings,
            "entities": {
                "accounts": sorted(accounts),
                "names": [],
                "amounts": amounts,
                "timestamps": sorted(timestamps),
                "locations": [],
            },
            "risk_signals": risk_signals,
            "evidence_alignment": [],
            "summary": "已完成 Excel 表头映射，并将全量交易数据转换为旧版混合模型可用结构。",
        }

    def _normalize_non_spreadsheet_bundle(self, bundle: dict, parser_summary: AnalysisParserSummary) -> dict:
        fallback = self._heuristic_normalize(bundle)
        if not settings.analysis_enable_llm_normalization:
            return fallback
        if not settings.analysis_llm_api_key:
            parser_summary.warnings.append("未配置 SiliconFlow API Key，非表格模态结构化已退回规则模式。")
            return fallback

        try:
            response_text = self._call_chat_completion(self._build_normalize_prompt(bundle))
            normalized = self._extract_json_object(response_text)
            return self._sanitize_llm_normalized_payload(normalized, fallback)
        except Exception as exc:
            parser_summary.warnings.append(f"非表格模态 LLM 结构化失败，已退回规则模式：{exc}")
            return fallback

    def _merge_normalized_payloads(self, spreadsheet_payload: dict, evidence_payload: dict) -> dict:
        spreadsheet_payload = self._sanitize_llm_normalized_payload(
            spreadsheet_payload,
            self._heuristic_normalize({"plain_texts": [], "documents_markdown": [], "structured_tables": [], "documents_json": []}),
        )
        evidence_payload = self._sanitize_llm_normalized_payload(
            evidence_payload,
            self._heuristic_normalize({"plain_texts": [], "documents_markdown": [], "structured_tables": [], "documents_json": []}),
        )
        accounts = set(self._stringify(item).strip() for item in spreadsheet_payload.get("entities", {}).get("accounts", []))
        accounts.update(self._stringify(item).strip() for item in evidence_payload.get("entities", {}).get("accounts", []))
        amounts = [self._safe_float(item) for item in spreadsheet_payload.get("entities", {}).get("amounts", [])]
        amounts.extend(self._safe_float(item) for item in evidence_payload.get("entities", {}).get("amounts", []))
        timestamps = set(self._stringify(item).strip() for item in spreadsheet_payload.get("entities", {}).get("timestamps", []))
        timestamps.update(self._stringify(item).strip() for item in evidence_payload.get("entities", {}).get("timestamps", []))

        risk_signals: list[str] = []
        for payload in (spreadsheet_payload, evidence_payload):
            for item in payload.get("risk_signals", []):
                text = self._stringify(item).strip()
                if text and text not in risk_signals:
                    risk_signals.append(text)

        summary_parts = [
            self._stringify(spreadsheet_payload.get("summary")).strip(),
            self._stringify(evidence_payload.get("summary")).strip(),
        ]
        summary = " ".join(part for part in summary_parts if part).strip()
        if not summary:
            summary = "已完成多模态证据归一化整理，可进入风险判定阶段。"

        return {
            "transaction_candidates": [
                *spreadsheet_payload.get("transaction_candidates", []),
                *evidence_payload.get("transaction_candidates", []),
            ],
            "standardized_transactions": [
                *spreadsheet_payload.get("standardized_transactions", []),
                *evidence_payload.get("standardized_transactions", []),
            ],
            "spreadsheet_mappings": spreadsheet_payload.get("spreadsheet_mappings", []),
            "entities": {
                "accounts": sorted(account for account in accounts if account),
                "names": [],
                "amounts": [amount for amount in amounts if amount],
                "timestamps": sorted(timestamp for timestamp in timestamps if timestamp),
                "locations": [],
            },
            "risk_signals": risk_signals,
            "evidence_alignment": [
                *spreadsheet_payload.get("evidence_alignment", []),
                *evidence_payload.get("evidence_alignment", []),
            ],
            "summary": summary,
        }

    def _infer_spreadsheet_mapping(self, table: dict, parser_summary: AnalysisParserSummary) -> dict:
        fallback = self._fallback_spreadsheet_mapping(table)
        if not settings.analysis_enable_llm_normalization or not settings.analysis_llm_api_key:
            return fallback

        try:
            response_text = self._call_chat_completion(self._build_spreadsheet_mapping_prompt(table))
            payload = self._extract_json_object(response_text)
            payload_column_mapping = (
                payload.get("column_mapping") if isinstance(payload.get("column_mapping"), dict) else {}
            )
            payload_static_mapping = (
                payload.get("static_mapping") if isinstance(payload.get("static_mapping"), dict) else {}
            )
            payload_column_mapping = self._normalize_mapping_direction(
                payload_column_mapping,
                allowed_targets=set(STANDARD_TRANSACTION_FIELDS) - {"xb", "年龄"},
                table_columns={self._stringify(column).strip() for column in table.get("columns", [])},
            )
            payload_static_mapping = self._normalize_mapping_direction(
                payload_static_mapping,
                allowed_targets={"xb", "年龄"},
                table_columns={self._stringify(column).strip() for column in table.get("columns", [])},
            )
            payload = {
                "table_kind": self._normalize_spreadsheet_table_kind(payload.get("table_kind"), fallback["table_kind"]),
                "column_mapping": payload_column_mapping,
                "static_mapping": payload_static_mapping,
            }
            if not {"zhdh", "dfzh", "jdbj", "jyje"} <= set(payload["column_mapping"]):
                return fallback
            return payload
        except Exception as exc:
            parser_summary.warnings.append(
                f"{self._stringify(table.get('sheet_name')) or '工作表'} 的表头映射退回规则模式：{exc}"
            )
            return fallback

    def _sanitize_llm_normalized_payload(self, payload: dict, fallback: dict) -> dict:
        if not isinstance(payload, dict):
            return fallback

        normalized = dict(payload)
        normalized["summary"] = self._stringify(normalized.get("summary")).strip() or fallback.get("summary", "")
        normalized["transaction_candidates"] = self._coerce_transaction_candidates(
            normalized.get("transaction_candidates"),
            fallback.get("transaction_candidates", []),
        )
        normalized["standardized_transactions"] = self._coerce_standardized_transactions(
            normalized.get("standardized_transactions"),
            fallback.get("standardized_transactions", []),
        )
        normalized["entities"] = self._coerce_entities(
            normalized.get("entities"),
            fallback.get("entities", {}),
        )
        normalized["risk_signals"] = self._coerce_string_list(
            normalized.get("risk_signals"),
            fallback.get("risk_signals", []),
        )
        normalized["evidence_alignment"] = (
            normalized.get("evidence_alignment")
            if isinstance(normalized.get("evidence_alignment"), list)
            else fallback.get("evidence_alignment", [])
        )
        return normalized

    def _coerce_transaction_candidates(self, value: object, fallback: list[dict]) -> list[dict]:
        if not isinstance(value, list):
            return list(fallback)
        candidates: list[dict] = []
        for item in value:
            if not isinstance(item, dict):
                continue
            row = {
                "payer_account": self._stringify(item.get("payer_account")).strip(),
                "receiver_account": self._stringify(item.get("receiver_account")).strip(),
                "amount": self._safe_float(item.get("amount")),
                "timestamp": self._stringify(item.get("timestamp")).strip(),
                "channel": self._stringify(item.get("channel")).strip(),
                "description": self._stringify(item.get("description")).strip(),
                "evidence_refs": item.get("evidence_refs") if isinstance(item.get("evidence_refs"), list) else [],
            }
            if not any(
                [
                    row["payer_account"],
                    row["receiver_account"],
                    row["timestamp"],
                    row["channel"],
                    row["description"],
                    row["amount"] > 0,
                ]
            ):
                continue
            candidates.append(row)
        return candidates if candidates else list(fallback)

    def _coerce_standardized_transactions(self, value: object, fallback: list[dict]) -> list[dict]:
        if not isinstance(value, list):
            return list(fallback)
        rows: list[dict] = []
        for item in value:
            if not isinstance(item, dict):
                continue
            row = {
                "jylsxh": self._stringify(item.get("jylsxh")).strip(),
                "zhdh": self._stringify(item.get("zhdh")).strip(),
                "dfzh": self._stringify(item.get("dfzh")).strip(),
                "jdbj": self._normalize_direction_value(item.get("jdbj")),
                "jyje": self._safe_float(item.get("jyje")),
                "zhye": self._safe_float(item.get("zhye")),
                "dfhh": self._stringify(item.get("dfhh")).strip(),
                "jyrq": self._normalize_date_value(item.get("jyrq")),
                "jysj": self._normalize_time_value(item.get("jysj")),
                "jyqd": self._stringify(item.get("jyqd")).strip(),
                "dfmccd": self._safe_float(item.get("dfmccd")),
                "xb": self._normalize_gender_value(item.get("xb")),
                "年龄": self._safe_float(item.get("年龄")),
            }
            if not any(
                [
                    row["jylsxh"],
                    row["zhdh"],
                    row["dfzh"],
                    row["dfhh"],
                    row["jyrq"],
                    row["jysj"],
                    row["jyqd"],
                    row["jyje"] > 0,
                    row["zhye"] > 0,
                ]
            ):
                continue
            rows.append(row)
        return rows if rows else list(fallback)

    def _coerce_entities(self, value: object, fallback: dict) -> dict:
        if not isinstance(value, dict):
            return dict(fallback)
        accounts = [self._stringify(item).strip() for item in value.get("accounts", []) if self._stringify(item).strip()]
        amounts = [self._safe_float(item) for item in value.get("amounts", []) if self._safe_float(item)]
        timestamps = [self._stringify(item).strip() for item in value.get("timestamps", []) if self._stringify(item).strip()]
        return {
            "accounts": accounts or list((fallback.get("accounts") or [])),
            "names": value.get("names") if isinstance(value.get("names"), list) else list((fallback.get("names") or [])),
            "amounts": amounts or list((fallback.get("amounts") or [])),
            "timestamps": timestamps or list((fallback.get("timestamps") or [])),
            "locations": value.get("locations") if isinstance(value.get("locations"), list) else list((fallback.get("locations") or [])),
        }

    def _coerce_string_list(self, value: object, fallback: list[str]) -> list[str]:
        if not isinstance(value, list):
            return list(fallback)
        items = [self._stringify(item).strip() for item in value if self._stringify(item).strip()]
        return items if items else list(fallback)

    def _fallback_spreadsheet_mapping(self, table: dict) -> dict:
        columns = [self._stringify(column).strip() for column in table.get("columns", [])]
        column_mapping: dict[str, str] = {}
        static_mapping: dict[str, str] = {}
        for target_field, keywords in SPREADSHEET_FIELD_KEYWORDS.items():
            matched = self._find_matching_key(columns, keywords)
            if matched:
                if target_field in {"xb", "年龄"}:
                    static_mapping[target_field] = matched
                else:
                    column_mapping[target_field] = matched
        return {
            "table_kind": "transaction" if {"zhdh", "dfzh", "jdbj", "jyje"} <= set(column_mapping) else "unknown",
            "column_mapping": column_mapping,
            "static_mapping": static_mapping,
        }

    def _apply_spreadsheet_mapping(self, table: dict, mapping: dict) -> list[dict]:
        column_mapping = mapping.get("column_mapping") if isinstance(mapping.get("column_mapping"), dict) else {}
        static_mapping = mapping.get("static_mapping") if isinstance(mapping.get("static_mapping"), dict) else {}
        if not {"zhdh", "dfzh", "jdbj", "jyje"} <= set(column_mapping):
            return []

        standardized_rows: list[dict] = []
        for index, row in enumerate(table.get("rows", []), start=1):
            if not isinstance(row, dict):
                continue
            normalized_row = {field: None for field in STANDARD_TRANSACTION_FIELDS}
            for target_field, source_field in column_mapping.items():
                normalized_row[target_field] = row.get(source_field)
            for target_field, source_field in static_mapping.items():
                normalized_row[target_field] = row.get(source_field)

            normalized_row["jylsxh"] = self._stringify(normalized_row.get("jylsxh")).strip() or f"{table.get('sheet_name', 'sheet')}-{index}"
            normalized_row["zhdh"] = self._stringify(normalized_row.get("zhdh")).strip()
            normalized_row["dfzh"] = self._stringify(normalized_row.get("dfzh")).strip()
            normalized_row["jdbj"] = self._normalize_direction_value(normalized_row.get("jdbj"))
            normalized_row["jyje"] = self._safe_float(normalized_row.get("jyje"))
            normalized_row["zhye"] = self._safe_float(normalized_row.get("zhye"))
            normalized_row["dfhh"] = self._stringify(normalized_row.get("dfhh")).strip()
            normalized_row["jyrq"] = self._normalize_date_value(normalized_row.get("jyrq"))
            normalized_row["jysj"] = self._normalize_time_value(normalized_row.get("jysj"))
            normalized_row["jyqd"] = self._stringify(normalized_row.get("jyqd")).strip()
            normalized_row["dfmccd"] = self._safe_float(normalized_row.get("dfmccd"))
            normalized_row["xb"] = self._normalize_gender_value(normalized_row.get("xb"))
            normalized_row["年龄"] = self._safe_float(normalized_row.get("年龄"))
            if normalized_row["zhdh"] and normalized_row["dfzh"]:
                standardized_rows.append(normalized_row)
        return standardized_rows

    def _build_spreadsheet_mapping_prompt(self, table: dict) -> list[dict]:
        sample_rows = []
        for row in table.get("rows", [])[:5]:
            if isinstance(row, dict):
                sample_rows.append(self._to_json_safe(row))
        payload = {
            "sheet_name": table.get("sheet_name"),
            "columns": [self._stringify(column) for column in table.get("columns", [])],
            "sample_rows": sample_rows,
        }
        example_input = {
            "sheet_name": "Sheet1",
            "columns": [
                "交易流水序号",
                "账户代号",
                "对方账户",
                "借贷标记",
                "交易金额",
                "账户余额",
                "对方行号",
                "交易日期",
                "交易时间",
                "交易渠道",
                "性别",
                "年龄",
            ],
            "sample_rows": [
                {
                    "交易流水序号": "96DE006FD8F51A8A0250C1D37DB2DCFF",
                    "账户代号": "3242244504235523",
                    "对方账户": "BA8C172953D66632",
                    "借贷标记": "是",
                    "交易金额": 349.75,
                    "账户余额": 356.86,
                    "对方行号": "D41D8CD9",
                    "交易日期": "2020-03-01 00:00:00",
                    "交易时间": "04:41:09",
                    "交易渠道": "757B505C",
                    "性别": "男",
                    "年龄": 37,
                }
            ],
        }
        example_output = {
            "table_kind": "transaction",
            "column_mapping": {
                "jylsxh": "交易流水序号",
                "zhdh": "账户代号",
                "dfzh": "对方账户",
                "jdbj": "借贷标记",
                "jyje": "交易金额",
                "zhye": "账户余额",
                "dfhh": "对方行号",
                "jyrq": "交易日期",
                "jysj": "交易时间",
                "jyqd": "交易渠道",
            },
            "static_mapping": {
                "xb": "性别",
                "年龄": "年龄",
            },
        }
        return [
            {
                "role": "system",
                "content": (
                    "你是金融表格字段映射助手。"
                    "请根据表头和前几行样本，识别交易表字段映射。"
                    "只输出合法 JSON，字段必须包含 table_kind, column_mapping, static_mapping。"
                    "table_kind 只能是 transaction 或 unknown。"
                    "column_mapping 只映射交易字段，static_mapping 只映射 xb 和 年龄。"
                    "映射方向必须固定为：标准字段名 -> 原始列名。"
                    "标准字段只允许使用：jylsxh, zhdh, dfzh, jdbj, jyje, zhye, dfhh, jyrq, jysj, jyqd, dfmccd, xb, 年龄。"
                    "不要把原始列名写成 key，不要输出解释，不要输出多余字段。"
                ),
            },
            {
                "role": "user",
                "content": (
                    "下面是一个示例。\n\n"
                    f"输入：{json.dumps(example_input, ensure_ascii=False)}\n\n"
                    f"输出：{json.dumps(example_output, ensure_ascii=False)}"
                ),
            },
            {
                "role": "user",
                "content": json.dumps(payload, ensure_ascii=False),
            },
        ]

    def _normalize_mapping_direction(
        self,
        mapping: dict[str, str],
        *,
        allowed_targets: set[str],
        table_columns: set[str],
    ) -> dict[str, str]:
        normalized: dict[str, str] = {}
        for key, value in mapping.items():
            key_text = self._stringify(key).strip()
            value_text = self._stringify(value).strip()
            if key_text in allowed_targets and value_text in table_columns:
                normalized[key_text] = value_text
            elif value_text in allowed_targets and key_text in table_columns:
                normalized[value_text] = key_text
        return normalized

    def _normalize_spreadsheet_table_kind(self, value: object, fallback: str) -> str:
        normalized = self._stringify(value).strip().lower()
        mapping = {
            "transaction": "transaction",
            "unknown": "unknown",
            "交易表": "transaction",
            "交易": "transaction",
            "表格": "transaction",
            "未知": "unknown",
        }
        return mapping.get(normalized, fallback)

    def _build_hybrid_result_payload(self, normalized: dict, hybrid_result: dict) -> dict:
        accounts = hybrid_result.get("accounts") if isinstance(hybrid_result.get("accounts"), list) else []
        overall = hybrid_result.get("overall") if isinstance(hybrid_result.get("overall"), dict) else {}
        risk_level = self._normalize_risk_level_value(overall.get("risk_level"), "medium")
        confidence = self._safe_float(overall.get("confidence")) or 0.0

        account_scores = []
        for account_item in accounts:
            if not isinstance(account_item, dict):
                continue
            probability = self._safe_float(account_item.get("probability"))
            account_scores.append(
                {
                    "account": self._stringify(account_item.get("account")).strip() or "待确认账户",
                    "prediction": int(self._safe_float(account_item.get("prediction"))),
                    "probability": round(probability, 6),
                    "gru_probability": round(self._safe_float(account_item.get("gru_probability")), 6),
                    "xgb_probability": round(self._safe_float(account_item.get("xgb_probability")), 6),
                    "confidence_label": self._stringify(account_item.get("confidence_label")).strip()
                    or self._confidence_label_from_probability(probability),
                    "risk_level": self._normalize_risk_level_value(account_item.get("risk_level"), risk_level),
                }
            )

        link_path = []
        for account_item in account_scores[:5]:
            link_path.append(
                AnalysisRiskNode(
                    account=account_item["account"],
                    risk_level=account_item["risk_level"],
                    action="复核" if account_item["risk_level"] != "low" else "记录",
                ).model_dump()
            )
        if not link_path:
            link_path = [AnalysisRiskNode(account="待补充证据", risk_level=risk_level, action="复核").model_dump()]

        top_account = account_scores[0]["account"] if account_scores else "待确认账户"
        narrative = (
            f"结构化交易数据已进入旧版混合模型，当前最高风险账户为 {top_account}，"
            f"综合风险等级为 {self._display_risk_level(risk_level)}。"
        )
        return {
            "risk_level": risk_level,
            "confidence": confidence,
            "model_source": self._stringify(hybrid_result.get("model_source")) or "legacy-hybrid-gru-xgb-meta",
            "narrative": narrative,
            "suggested_actions": self._build_actions(risk_level),
            "link_path": link_path,
            "account_scores": account_scores,
            "normalized_summary": normalized.get("summary", ""),
            "risk_signals": normalized.get("risk_signals", []),
            "transaction_candidates": normalized.get("transaction_candidates", []),
        }

    def _normalize_bundle(self, bundle: dict, parser_summary: AnalysisParserSummary) -> dict:
        spreadsheet_normalized = self._normalize_spreadsheets(bundle.get("structured_tables", []), parser_summary)
        evidence_bundle = {
            **bundle,
            "structured_tables": [],
        }
        evidence_normalized = self._normalize_non_spreadsheet_bundle(evidence_bundle, parser_summary)
        return self._merge_normalized_payloads(spreadsheet_normalized, evidence_normalized)

    def _inject_focus_history(self, normalized: dict, current_job_id: str) -> dict:
        targets = [
            item
            for item in focus_repository.list_targets()
            if bool(item.get("is_seed")) and self._stringify(item.get("account")).strip()
        ]
        if not targets:
            return normalized

        history_jobs = [
            job
            for job in repository.list_jobs(created_by=None)
            if job["status"] == "completed"
            and job["job_id"] != current_job_id
            and isinstance(job.get("normalized_json"), dict)
        ]
        if not history_jobs:
            return normalized

        normal_accounts = {
            self._stringify(item.get("account")).strip()
            for item in targets
            if item.get("mode") == "normal"
        }
        deep_seed_accounts = {
            self._stringify(item.get("account")).strip()
            for item in targets
            if item.get("mode") == "deep"
        }

        injected_transactions: list[dict] = []
        normal_transactions = self._collect_history_transactions(history_jobs, normal_accounts)
        injected_transactions.extend(normal_transactions)

        deep_seed_transactions = self._collect_history_transactions(history_jobs, deep_seed_accounts)
        deep_counterparts = self._collect_counterpart_accounts(deep_seed_transactions, deep_seed_accounts)
        deep_scope_accounts = deep_seed_accounts | deep_counterparts
        deep_transactions = self._collect_history_transactions(history_jobs, deep_scope_accounts)
        injected_transactions.extend(deep_transactions)

        if not injected_transactions:
            return normalized

        merged = dict(normalized)
        existing_transactions = list(normalized.get("standardized_transactions") or [])
        merged_transactions = self._append_unique_transactions(existing_transactions, injected_transactions)
        added_count = max(0, len(merged_transactions) - len(existing_transactions))
        if added_count <= 0:
            return normalized

        merged["standardized_transactions"] = merged_transactions

        existing_candidates = list(normalized.get("transaction_candidates") or [])
        focus_candidates = [self._build_candidate_from_transaction(item) for item in injected_transactions]
        merged["transaction_candidates"] = self._append_unique_candidates(existing_candidates, focus_candidates)

        entities = dict(normalized.get("entities") or {})
        entities["accounts"] = self._merge_string_values(
            entities.get("accounts"),
            self._extract_accounts_from_transactions(injected_transactions),
        )
        entities["amounts"] = self._merge_float_values(
            entities.get("amounts"),
            [
                self._safe_float(item.get("jyje"))
                for item in injected_transactions
                if self._safe_float(item.get("jyje"))
            ]
            + [
                self._safe_float(item.get("zhye"))
                for item in injected_transactions
                if self._safe_float(item.get("zhye"))
            ],
        )
        entities["timestamps"] = self._merge_string_values(
            entities.get("timestamps"),
            self._extract_timestamps_from_transactions(injected_transactions),
        )
        merged["entities"] = entities

        risk_signals = list(normalized.get("risk_signals") or [])
        signal_parts: list[str] = []
        if normal_transactions:
            signal_parts.append(f"普通关注历史日志 {len(normal_transactions)} 条")
        if deep_transactions:
            signal_parts.append(f"深度追踪历史日志 {len(deep_transactions)} 条")
        if signal_parts:
            merged["risk_signals"] = [f"已自动并入重点关注交易日志：{'，'.join(signal_parts)}", *risk_signals]

        evidence_alignment = list(normalized.get("evidence_alignment") or [])
        evidence_alignment.append(
            {
                "asset_id": "focus-history",
                "mapped_fields": list(STANDARD_TRANSACTION_FIELDS),
                "focus_accounts": sorted(normal_accounts | deep_seed_accounts),
                "deep_counterparts": sorted(deep_counterparts),
                "injected_transactions": added_count,
            }
        )
        merged["evidence_alignment"] = evidence_alignment

        summary = self._stringify(normalized.get("summary")).strip()
        focus_summary = f"已自动并入重点关注历史交易日志 {added_count} 条。"
        merged["summary"] = f"{summary} {focus_summary}".strip() if summary else focus_summary
        return merged

    def _collect_history_transactions(self, jobs: list[dict], accounts: set[str]) -> list[dict]:
        if not accounts:
            return []

        rows: list[dict] = []
        seen: set[tuple] = set()
        for job in jobs:
            normalized = job.get("normalized_json") or {}
            candidates = normalized.get("transaction_candidates") or []
            for row in normalized.get("standardized_transactions") or []:
                if not self._transaction_matches_accounts(row, accounts):
                    continue
                normalized_row = self._normalize_history_transaction_row(row)
                key = self._build_transaction_key(normalized_row)
                if key in seen:
                    continue
                seen.add(key)
                rows.append(normalized_row)

            if normalized.get("standardized_transactions"):
                continue

            for candidate in candidates:
                pseudo_row = self._candidate_to_transaction_row(candidate)
                if not self._transaction_matches_accounts(pseudo_row, accounts):
                    continue
                key = self._build_transaction_key(pseudo_row)
                if key in seen:
                    continue
                seen.add(key)
                rows.append(pseudo_row)
        return rows

    def _collect_counterpart_accounts(self, transactions: list[dict], seed_accounts: set[str]) -> set[str]:
        counterparts: set[str] = set()
        for row in transactions:
            payer = self._stringify(row.get("zhdh")).strip()
            receiver = self._stringify(row.get("dfzh")).strip()
            if payer in seed_accounts and receiver:
                counterparts.add(receiver)
            if receiver in seed_accounts and payer:
                counterparts.add(payer)
        return counterparts - seed_accounts

    def _transaction_matches_accounts(self, row: dict, accounts: set[str]) -> bool:
        payer = self._stringify(row.get("zhdh")).strip()
        receiver = self._stringify(row.get("dfzh")).strip()
        return payer in accounts or receiver in accounts

    def _normalize_history_transaction_row(self, row: dict) -> dict:
        return {
            "jylsxh": self._stringify(row.get("jylsxh")).strip(),
            "zhdh": self._stringify(row.get("zhdh")).strip(),
            "dfzh": self._stringify(row.get("dfzh")).strip(),
            "jdbj": self._normalize_direction_value(row.get("jdbj")),
            "jyje": self._safe_float(row.get("jyje")),
            "zhye": self._safe_float(row.get("zhye")),
            "dfhh": self._stringify(row.get("dfhh")).strip(),
            "jyrq": self._normalize_date_value(row.get("jyrq")),
            "jysj": self._normalize_time_value(row.get("jysj")),
            "jyqd": self._stringify(row.get("jyqd")).strip(),
            "dfmccd": self._safe_float(row.get("dfmccd")),
            "xb": self._normalize_gender_value(row.get("xb")),
            "年龄": self._safe_float(row.get("年龄")),
        }

    def _candidate_to_transaction_row(self, row: dict) -> dict:
        timestamp = self._stringify(row.get("timestamp")).strip()
        date_part, time_part = "", ""
        if timestamp:
            pieces = timestamp.replace("T", " ").split(" ", 1)
            date_part = self._normalize_date_value(pieces[0])
            time_part = self._normalize_time_value(pieces[1] if len(pieces) > 1 else "")

        return {
            "jylsxh": "",
            "zhdh": self._stringify(row.get("payer_account")).strip(),
            "dfzh": self._stringify(row.get("receiver_account")).strip(),
            "jdbj": 1,
            "jyje": self._safe_float(row.get("amount")),
            "zhye": 0.0,
            "dfhh": "",
            "jyrq": date_part,
            "jysj": time_part,
            "jyqd": self._stringify(row.get("channel")).strip(),
            "dfmccd": 0.0,
            "xb": 0,
            "年龄": 0.0,
        }

    def _append_unique_transactions(self, base: list[dict], extras: list[dict]) -> list[dict]:
        merged = list(base)
        seen = {self._build_transaction_key(item) for item in merged}
        for item in extras:
            key = self._build_transaction_key(item)
            if key in seen:
                continue
            seen.add(key)
            merged.append(item)
        return merged

    def _append_unique_candidates(self, base: list[dict], extras: list[dict]) -> list[dict]:
        merged = list(base)
        seen = {self._build_candidate_key(item) for item in merged if isinstance(item, dict)}
        for item in extras:
            key = self._build_candidate_key(item)
            if key in seen:
                continue
            seen.add(key)
            merged.append(item)
        return merged

    def _build_transaction_key(self, row: dict) -> tuple:
        return (
            self._stringify(row.get("jylsxh")).strip(),
            self._stringify(row.get("zhdh")).strip(),
            self._stringify(row.get("dfzh")).strip(),
            self._normalize_date_value(row.get("jyrq")),
            self._normalize_time_value(row.get("jysj")),
            round(self._safe_float(row.get("jyje")), 6),
            round(self._safe_float(row.get("zhye")), 6),
            self._stringify(row.get("jyqd")).strip(),
        )

    def _build_candidate_key(self, row: dict) -> tuple:
        return (
            self._stringify(row.get("payer_account")).strip(),
            self._stringify(row.get("receiver_account")).strip(),
            round(self._safe_float(row.get("amount")), 6),
            self._stringify(row.get("timestamp")).strip(),
            self._stringify(row.get("channel")).strip(),
        )

    def _build_candidate_from_transaction(self, row: dict) -> dict:
        timestamp = " ".join(
            part
            for part in [
                self._normalize_date_value(row.get("jyrq")),
                self._normalize_time_value(row.get("jysj")),
            ]
            if part
        ).strip()
        return {
            "payer_account": self._stringify(row.get("zhdh")).strip(),
            "receiver_account": self._stringify(row.get("dfzh")).strip(),
            "amount": self._safe_float(row.get("jyje")),
            "timestamp": timestamp,
            "channel": self._stringify(row.get("jyqd")).strip() or "重点关注历史交易日志",
            "description": "重点关注历史交易日志",
            "evidence_refs": ["focus-history"],
        }

    def _extract_accounts_from_transactions(self, transactions: list[dict]) -> list[str]:
        accounts: list[str] = []
        seen: set[str] = set()
        for row in transactions:
            for value in (row.get("zhdh"), row.get("dfzh")):
                account = self._stringify(value).strip()
                if account and account not in seen:
                    seen.add(account)
                    accounts.append(account)
        return accounts

    def _extract_timestamps_from_transactions(self, transactions: list[dict]) -> list[str]:
        timestamps: list[str] = []
        seen: set[str] = set()
        for row in transactions:
            timestamp = " ".join(
                part
                for part in [
                    self._normalize_date_value(row.get("jyrq")),
                    self._normalize_time_value(row.get("jysj")),
                ]
                if part
            ).strip()
            if timestamp and timestamp not in seen:
                seen.add(timestamp)
                timestamps.append(timestamp)
        return timestamps

    def _merge_string_values(self, base: object, extras: list[str]) -> list[str]:
        merged: list[str] = []
        seen: set[str] = set()
        for item in list(base or []) + list(extras or []):
            value = self._stringify(item).strip()
            if value and value not in seen:
                seen.add(value)
                merged.append(value)
        return merged

    def _merge_float_values(self, base: object, extras: list[float]) -> list[float]:
        merged: list[float] = []
        seen: set[float] = set()
        for item in list(base or []) + list(extras or []):
            value = self._safe_float(item)
            if not value:
                continue
            normalized_value = round(value, 6)
            if normalized_value in seen:
                continue
            seen.add(normalized_value)
            merged.append(value)
        return merged

    def _analyze_bundle(self, normalized: dict, parser_summary: AnalysisParserSummary, job_dir: Path) -> dict:
        fallback = self._heuristic_analysis(normalized)
        standardized_transactions = normalized.get("standardized_transactions") or []
        if not standardized_transactions:
            parser_summary.warnings.append("未提取到可进入旧版混合模型的结构化交易数据，已退回规则分析。")
            return fallback

        try:
            hybrid_result = hybrid_adapter.predict(
                transactions=standardized_transactions,
                work_dir=job_dir / "_hybrid_runtime",
                inference_params=params_service.get_hybrid_inference_params(),
            )
            payload = self._build_hybrid_result_payload(normalized, hybrid_result)
            return self._normalize_result_payload(payload) or fallback
        except Exception as exc:
            parser_summary.warnings.append(f"旧版混合模型推理失败，已退回规则分析：{exc}")
            return fallback

    def _heuristic_normalize(self, bundle: dict) -> dict:
        accounts: set[str] = set()
        amounts: list[float] = []
        timestamps: set[str] = set()
        candidates: list[dict] = []
        risk_signals: list[str] = []

        for text_block in bundle["plain_texts"]:
            text = text_block["content"]
            accounts.update(re.findall(r"\b\d{8,32}\b", text))
            amounts.extend(float(item) for item in re.findall(r"\b\d+(?:\.\d{1,2})?\b", text)[:10])

        for document in bundle["documents_markdown"]:
            markdown = document["markdown"]
            accounts.update(re.findall(r"\b\d{8,32}\b", markdown))
            amounts.extend(float(item) for item in re.findall(r"\b\d+(?:\.\d{1,2})?\b", markdown)[:10])

        for table in bundle["structured_tables"]:
            columns = table["columns"]
            rows = table["rows"]
            payer_key = self._find_matching_key(columns, ("付款", "payer", "转出", "付款方"))
            receiver_key = self._find_matching_key(columns, ("收款", "receiver", "转入", "收款方"))
            amount_key = self._find_matching_key(columns, ("金额", "amount", "交易额"))
            time_key = self._find_matching_key(columns, ("时间", "日期", "timestamp"))
            for row in rows:
                payer_value = self._stringify(row.get(payer_key)) if payer_key else ""
                receiver_value = self._stringify(row.get(receiver_key)) if receiver_key else ""
                amount_value = self._safe_float(row.get(amount_key)) if amount_key else 0.0
                time_value = self._stringify(row.get(time_key)) if time_key else ""
                if payer_value:
                    accounts.add(payer_value)
                if receiver_value:
                    accounts.add(receiver_value)
                if amount_value:
                    amounts.append(amount_value)
                if time_value:
                    timestamps.add(time_value)
                if payer_value or receiver_value or amount_value:
                    candidates.append(
                        {
                            "payer_account": payer_value,
                            "receiver_account": receiver_value,
                            "amount": amount_value,
                            "timestamp": time_value,
                            "channel": "Excel结构化表格",
                            "description": f"{table['sheet_name']} 表单候选交易",
                            "evidence_refs": [table["sheet_name"]],
                        }
                    )

        if any(amount >= 50000 for amount in amounts):
            risk_signals.append("存在大额转账记录")
        if len(accounts) >= 3:
            risk_signals.append("检测到多账户关联")
        if bundle["documents_markdown"]:
            risk_signals.append("存在文档类多模态证据")

        return {
            "transaction_candidates": candidates,
            "standardized_transactions": [],
            "entities": {
                "accounts": sorted(accounts),
                "names": [],
                "amounts": amounts,
                "timestamps": sorted(timestamps),
                "locations": [],
            },
            "risk_signals": risk_signals,
            "evidence_alignment": [],
            "summary": "已完成多模态证据归一化整理，可进入风险判定阶段。",
        }

    def _heuristic_analysis(self, normalized: dict) -> dict:
        candidates = normalized.get("transaction_candidates", [])
        accounts = normalized.get("entities", {}).get("accounts", [])
        amounts = normalized.get("entities", {}).get("amounts", [])
        risk_signals = normalized.get("risk_signals", [])
        dynamic_thresholds = params_service.get_dynamic_thresholds()
        high_threshold = dynamic_thresholds["high_risk_threshold"]
        medium_threshold = dynamic_thresholds["medium_risk_threshold"]
        self_attention_enabled = dynamic_thresholds["self_attention_enabled"]
        adaptive_threshold_enabled = dynamic_thresholds["adaptive_threshold_enabled"]

        highest_amount = max((self._safe_float(value) for value in amounts), default=0.0)
        candidate_count = len(candidates)
        account_count = len(accounts)
        risk_signal_count = len(risk_signals)

        if adaptive_threshold_enabled:
            if risk_signal_count >= 3 or candidate_count >= 2:
                high_threshold = max(medium_threshold * 1.6, high_threshold * 0.72)
            if risk_signal_count >= 2 or account_count >= 2:
                medium_threshold = max(1.0, medium_threshold * 0.7)

        attention_score = 0.0
        if self_attention_enabled:
            attention_score += min(highest_amount / max(high_threshold, 1.0), 1.4) * 0.42
            attention_score += min(risk_signal_count / 3.0, 1.0) * 0.26
            attention_score += min(account_count / 4.0, 1.0) * 0.18
            attention_score += min(candidate_count / 3.0, 1.0) * 0.14

        if highest_amount >= high_threshold or risk_signal_count >= 3 or attention_score >= 0.86:
            risk_level = "high"
            confidence = min(0.95, 0.78 + min(attention_score, 1.0) * 0.14)
        elif highest_amount >= medium_threshold or account_count >= 2 or attention_score >= 0.56:
            risk_level = "medium"
            confidence = min(0.86, 0.64 + min(attention_score, 1.0) * 0.12)
        else:
            risk_level = "low"
            confidence = max(0.52, 0.58 + min(attention_score, 0.3) * 0.08)

        if candidates:
            primary = candidates[0]
            link_path = [
                AnalysisRiskNode(
                    account=primary.get("payer_account") or "待确认付款方",
                    risk_level="medium",
                    action="观察",
                ).model_dump(),
                AnalysisRiskNode(
                    account=primary.get("receiver_account") or "待确认收款方",
                    risk_level=risk_level,
                    action="复核" if risk_level != "low" else "记录",
                ).model_dump(),
            ]
        else:
            link_path = [
                AnalysisRiskNode(account=account, risk_level=risk_level, action="复核").model_dump()
                for account in accounts[:2]
            ]
        if not link_path:
            link_path = [AnalysisRiskNode(account="待补充证据", risk_level=risk_level, action="复核").model_dump()]

        return {
            "risk_level": risk_level,
            "confidence": round(confidence, 4),
            "model_source": "dynamic-heuristic",
            "narrative": "系统已基于多模态证据完成初步判定，建议结合报告中的关键证据进行人工复核。",
            "suggested_actions": self._build_actions(risk_level),
            "link_path": link_path,
            "normalized_summary": normalized.get("summary", ""),
            "risk_signals": risk_signals,
            "transaction_candidates": candidates,
        }

    def _render_report(self, *, job_id: str, assets: list[dict], normalized: dict, result: dict) -> Path:
        report_dir = self.storage_dir / job_id
        report_dir.mkdir(parents=True, exist_ok=True)
        report_path = report_dir / "report.html"
        report_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        background_uri = self._build_background_data_uri()
        user_info_html = self._build_report_user_info_table(normalized, result)
        prediction_html = self._build_report_prediction_table(result)
        link_analysis_text = self._build_report_link_analysis_text(assets, normalized, result)
        link_graph_uri = self._build_link_analysis_svg(result)
        chart_rows = self._build_report_chart_rows(assets, normalized, result)

        html_content = f"""
<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <title>{REPORT_TITLE}</title>
  <style>
    @page {{
      size: A4;
      margin: 0mm;
    }}
    html {{
      font-family: "Microsoft YaHei", "SimHei", Arial, sans-serif;
      color: #333;
      margin: 0;
      padding: 0;
      width: 100%;
      height: 100%;
      box-sizing: border-box;
      -webkit-print-color-adjust: exact !important;
      print-color-adjust: exact !important;
    }}
    body {{
      margin: 0;
      padding: 0;
      width: 100%;
      min-height: 100%;
      box-sizing: border-box;
      background-image: url('{background_uri}');
      background-repeat: no-repeat;
      background-position: center center;
      background-size: 100% 100%;
      -webkit-print-color-adjust: exact !important;
      print-color-adjust: exact !important;
    }}
    .top-spacer {{
      height: 25mm;
      width: 100%;
    }}
    .page-content {{
      margin: 0 15mm 12mm;
      padding: 1px;
      box-sizing: border-box;
      position: relative;
      z-index: 1;
    }}
    h1, h2 {{
      text-align: center;
      color: #2c3e50;
    }}
    h1 {{
      font-size: 28px;
      margin-top: 0;
      padding-top: 10px;
    }}
    .subtitle {{
      font-size: 16px;
      text-align: center;
      margin-top: 10px;
      margin-bottom: 30px;
      color: #555;
    }}
    h2 {{
      font-size: 22px;
      margin-top: 40px;
      border-bottom: 2px solid #ccc;
      padding-bottom: 5px;
    }}
    .data-table {{
      width: 100%;
      border-collapse: collapse;
      margin-bottom: 30px;
      font-size: 14px;
      background-color: #fff;
    }}
    .data-table th, .data-table td {{
      border: 1px solid #ccc;
      padding: 8px;
      text-align: center;
    }}
    .user-info-table th {{
      background-color: #2980b9;
      color: white;
    }}
    .data-table th {{
      background-color: #3498db;
      color: white;
    }}
    .img-table {{
      width: 100%;
      border-collapse: separate;
      border-spacing: 15px 5px;
      margin-bottom: 40px;
    }}
    .chart-td {{
      width: 50%;
      vertical-align: top;
      padding: 0;
    }}
    .chart-card {{
      background-color: #fff;
      border-radius: 8px;
      padding: 12px 15px;
      box-shadow: 0 2px 8px rgba(0,0,0,0.10);
      text-align: center;
      height: 320px;
      box-sizing: border-box;
      display: flex;
      flex-direction: column;
      justify-content: flex-start;
      overflow: hidden;
    }}
    .chart-card h3 {{
      font-size: 18px;
      margin-top: 0;
      margin-bottom: -15px;
      color: #2c3e50;
      height: 44px;
      line-height: 22px;
      flex-shrink: 0;
    }}
    .chart-card img {{
      max-width: 100%;
      max-height: calc(100% - 44px - 12px - 2px);
      object-fit: contain;
      border: 1px solid #ccc;
      border-radius: 4px;
      margin: 0 auto;
      flex-grow: 1;
      display: block;
      background: #fff;
    }}
    .chart-card-full {{
      background-color: #fff;
      border-radius: 8px;
      padding: 15px;
      box-shadow: 0 2px 8px rgba(0,0,0,0.10);
      text-align: center;
      margin-bottom: 25px;
      page-break-inside: avoid;
    }}
    .chart-card-full h3 {{
      font-size: 20px;
      margin-top: 0;
      margin-bottom: 15px;
    }}
    .chart-card-full img {{
      width: 100%;
      max-height: 340px;
      object-fit: contain;
      border: 1px solid #ddd;
      border-radius: 4px;
      background: #fff;
    }}
    .analysis-text-box {{
      background-color: #f8f9fa;
      border-left: 5px solid #3498db;
      padding: 15px 20px;
      margin-top: 35px;
      font-size: 15px;
      line-height: 1.6;
      text-align: justify;
    }}
    .analysis-text-box p {{
      margin-bottom: 12px;
    }}
    .analysis-text-box p:last-child {{
      margin-bottom: 0;
    }}
  </style>
</head>
<body>
  <div class="page-content">
    <div class="top-spacer"></div>
    <h1>{REPORT_TITLE}</h1>
    <div class="subtitle">
      当前使用模型版本：1.56&nbsp;&nbsp;&nbsp;&nbsp;
      当前使用参数版本：2.5&nbsp;&nbsp;&nbsp;&nbsp;
      生成报告时间：{report_time}
    </div>
    <h2>一、交易账户静态信息</h2>
    {user_info_html}
    <h2>二、预测结果</h2>
    {prediction_html}
    <h2>三、链路分析</h2>
    <div class="chart-card-full">
      <h3>金融交易链路分析图</h3>
      <img src="{link_graph_uri}" alt="金融交易链路分析图" />
    </div>
    <div class="analysis-text-box">
      {link_analysis_text}
    </div>
    <h2>四、分析图表</h2>
    <table class="img-table">
      {chart_rows}
    </table>
  </div>
</body>
</html>
"""
        report_path.write_text(html_content.strip(), encoding="utf-8")
        self._write_pdf_report(report_path, report_dir / "report.pdf")
        return report_path

    def _ensure_report_files(self, job_id: str, job: dict) -> Path:
        report_path = Path(job["report_path"]) if job.get("report_path") else self.storage_dir / job_id / "report.html"
        pdf_path = report_path.with_suffix(".pdf")
        if report_path.exists() and pdf_path.exists():
            return report_path

        normalized = job.get("normalized_json")
        result = job.get("result_json")
        if not normalized or not result:
            raise AppError("分析报告尚未生成", status_code=404, code="ANALYSIS_REPORT_NOT_FOUND")

        assets = repository.list_assets(job_id)
        report_path = self._render_report(job_id=job_id, assets=assets, normalized=normalized, result=result)
        repository.update_job(
            job_id,
            {
                "report_path": str(report_path),
                "report_title": REPORT_TITLE,
            },
        )
        return report_path

    def _build_background_data_uri(self) -> str:
        legacy_root = Path(settings.legacy_result_dir).parent
        for background_path, mime_type in (
            (legacy_root / "back.png", "image/png"),
            (legacy_root / "back.jpg", "image/jpeg"),
            (legacy_root / "back.jpeg", "image/jpeg"),
        ):
            if background_path.exists():
                return self._file_to_data_uri(background_path, mime_type)
        return "none"

    def _build_report_user_info_table(self, normalized: dict, result: dict) -> str:
        rows = self._extract_report_user_rows(normalized, result)
        body = "".join(
            "<tr>"
            f"<td>{html.escape(item['account'])}</td>"
            f"<td>{html.escape(item['gender'])}</td>"
            f"<td>{html.escape(item['age'])}</td>"
            "</tr>"
            for item in rows
        )
        return (
            '<table class="data-table user-info-table">'
            "<thead><tr><th>转账账户</th><th>性别</th><th>年龄</th></tr></thead>"
            f"<tbody>{body}</tbody></table>"
        )

    def _build_report_prediction_table(self, result: dict) -> str:
        score_rows = self._get_report_account_scores(result)
        if not score_rows:
            probability = float(result.get("confidence", 0.0))
            prediction = 1 if self._normalize_risk_level_value(result.get("risk_level"), "low") == "high" else 0
            score_rows = [
                {
                    "prediction": prediction,
                    "probability": probability,
                    "gru_probability": probability,
                    "xgb_probability": probability,
                    "confidence_label": self._confidence_label_from_probability(probability),
                }
            ]

        body_rows: list[str] = []
        for item in score_rows:
            confidence_label = self._stringify(item.get("confidence_label")).strip() or self._confidence_label_from_probability(
                self._safe_float(item.get("probability"))
            )
            row_style = ""
            if "高置信度" in confidence_label:
                row_style = ' style="background-color: #FFC0CB;"'
            elif "中置信度" in confidence_label:
                row_style = ' style="background-color: #FFFF99;"'
            body_rows.append(
                f"<tr{row_style}>"
                f"<td>{html.escape(self._prediction_label(int(self._safe_float(item.get('prediction')))))}</td>"
                f"<td>{self._safe_float(item.get('probability')):.6f}</td>"
                f"<td>{self._safe_float(item.get('gru_probability')):.6f}</td>"
                f"<td>{self._safe_float(item.get('xgb_probability')):.6f}</td>"
                f"<td>{html.escape(confidence_label)}</td>"
                "</tr>"
            )

        return (
            '<table class="data-table"><tbody>'
            "<tr><th>预测结果</th><th>AT-GNN模型总概率</th><th>传递模块概率</th><th>图核模块概率</th><th>置信度评估</th></tr>"
            f"{''.join(body_rows)}"
            "</tbody></table>"
        )

    def _build_report_link_analysis_text(self, assets: list[dict], normalized: dict, result: dict) -> str:
        candidates = result.get("transaction_candidates", [])
        accounts = [self._stringify(item.get("account")).strip() for item in result.get("link_path", []) if self._stringify(item.get("account")).strip()]
        amounts = [self._safe_float(item.get("amount")) for item in candidates]
        valid_amounts = [amount for amount in amounts if amount > 0]
        highest_amount = max(valid_amounts, default=0.0)
        image_count = len([asset for asset in assets if self._stringify(asset.get("mime_type")).startswith("image/")])
        document_count = len([asset for asset in assets if asset.get("asset_type") == "document"])
        spreadsheet_count = len([asset for asset in assets if asset.get("asset_type") == "spreadsheet"])
        strongest_account = accounts[0] if accounts else "待确认账户"
        counterpart = accounts[1] if len(accounts) > 1 else "待确认对手账户"
        risk_signals = result.get("risk_signals", [])

        paragraphs = [
            (
                f"<p><strong>总体概览：</strong>本次多模态分析共覆盖 <strong>{len(assets)}</strong> 份输入资产，"
                f"其中包含 <strong>{document_count}</strong> 份文档/图片证据、"
                f"<strong>{spreadsheet_count}</strong> 份结构化表格，识别出 <strong>{len(candidates)}</strong> 条候选交易线索。</p>"
            ),
            (
                f"<p><strong>关键节点分析：</strong>账户 <strong>{html.escape(strongest_account)}</strong> 为当前链路中的重点关注节点，"
                f"与账户 <strong>{html.escape(counterpart)}</strong> 构成了最主要的资金关联路径，"
                f"整体风险等级判定为 <strong>{html.escape(self._display_risk_level(result.get('risk_level', 'low')))}</strong>。</p>"
            ),
            (
                f"<p><strong>交易模式分析：</strong>当前识别到的最高交易金额为 <strong>{highest_amount:.2f}</strong>，"
                f"已提取 <strong>{len(risk_signals)}</strong> 项风险信号。"
                f"{'已纳入图片凭证交叉核验。' if image_count else '本次未上传图片凭证。'}</p>"
            ),
            (
                f"<p><strong>结论与建议：</strong>{html.escape(self._stringify(result.get('narrative')))}"
                "建议结合下方图表与证据预览，对关键账户、关键金额与上传凭证进行进一步复核。</p>"
            ),
        ]
        return "".join(paragraphs)

    def _build_report_chart_rows(self, assets: list[dict], normalized: dict, result: dict) -> str:
        cards: list[tuple[str, str]] = [
            ("预测结果分布图", self._build_prediction_distribution_svg(result)),
            ("样本置信度分布", self._build_confidence_distribution_svg(result)),
            ("预测概率分布图", self._build_probability_distribution_svg(result)),
            ("传递模块与图核模块概率对比", self._build_model_comparison_svg(result)),
            ("高风险样本 TOP3 特征均值", self._build_high_risk_features_svg(normalized, result)),
            ("预测概率的箱线图与小提琴图", self._build_probability_violin_boxplot_svg(result)),
            ("模型概率之间的相关性热力图", self._build_correlation_heatmap_svg(result)),
            ("置信度与预测标签的关系", self._build_confidence_vs_prediction_svg(result)),
        ]

        rows: list[str] = []
        for index in range(0, len(cards), 2):
            left_title, left_src = cards[index]
            right_title, right_src = cards[index + 1]
            rows.append(
                "<tr>"
                f"{self._build_chart_cell(left_title, left_src)}"
                f"{self._build_chart_cell(right_title, right_src)}"
                "</tr>"
            )
        return "".join(rows)

    def _build_chart_cell(self, title: str, src: str) -> str:
        return (
            '<td class="chart-td"><div class="chart-card">'
            f"<h3>{html.escape(title)}</h3>"
            f'<img src="{src}" alt="{html.escape(title)}" />'
            "</div></td>"
        )

    def _extract_report_user_rows(self, normalized: dict, result: dict) -> list[dict[str, str]]:
        account_rows = self._get_report_account_scores(result)
        account_meta: dict[str, dict[str, str]] = {}
        for transaction in normalized.get("standardized_transactions", []) or []:
            if not isinstance(transaction, dict):
                continue
            account = self._stringify(transaction.get("zhdh")).strip()
            if not account or account in account_meta:
                continue
            account_meta[account] = {
                "gender": self._display_gender(transaction.get("xb")),
                "age": self._display_age(transaction.get("年龄")),
            }

        rows: list[dict[str, str]] = []
        seen_accounts: set[str] = set()
        for item in account_rows:
            account = self._stringify(item.get("account")).strip()
            if not account or account in seen_accounts:
                continue
            seen_accounts.add(account)
            meta = account_meta.get(account, {})
            rows.append(
                {
                    "account": account,
                    "gender": meta.get("gender", "-"),
                    "age": meta.get("age", "-"),
                }
            )

        for account in normalized.get("entities", {}).get("accounts", []) or []:
            account_text = self._stringify(account).strip()
            if not account_text or account_text in seen_accounts:
                continue
            seen_accounts.add(account_text)
            meta = account_meta.get(account_text, {})
            rows.append(
                {
                    "account": account_text,
                    "gender": meta.get("gender", "-"),
                    "age": meta.get("age", "-"),
                }
            )
            if len(rows) >= 8:
                break

        if not rows:
            rows.append({"account": "待补充证据", "gender": "-", "age": "-"})
        return rows

    def _get_report_account_scores(self, result: dict) -> list[dict]:
        raw_rows = result.get("account_scores")
        normalized_rows: list[dict] = []
        if isinstance(raw_rows, list):
            for item in raw_rows:
                if not isinstance(item, dict):
                    continue
                normalized_rows.append(
                    {
                        "account": self._stringify(item.get("account")).strip(),
                        "prediction": int(self._safe_float(item.get("prediction"))),
                        "probability": self._safe_float(item.get("probability")),
                        "gru_probability": self._safe_float(item.get("gru_probability")),
                        "xgb_probability": self._safe_float(item.get("xgb_probability")),
                        "confidence_label": self._stringify(item.get("confidence_label")).strip(),
                    }
                )
        normalized_rows = [item for item in normalized_rows if item.get("account")]
        normalized_rows.sort(key=lambda item: item.get("probability", 0.0), reverse=True)
        return normalized_rows

    def _display_gender(self, value: object) -> str:
        normalized = self._normalize_gender_value(value)
        if normalized == 1:
            return "男"
        if normalized == 0 and self._stringify(value).strip():
            return "女"
        return "-"

    def _display_age(self, value: object) -> str:
        age = int(round(self._safe_float(value)))
        return str(age) if age > 0 else "-"

    def _confidence_label_from_probability(self, probability: float) -> str:
        if probability > 0.7:
            return "高置信度"
        if probability > 0.3:
            return "中置信度"
        return "低置信度"

    def _prediction_label(self, prediction: int) -> str:
        return "欺诈账户" if prediction == 1 else "非欺诈账户"

    def _build_prediction_distribution_svg(self, result: dict) -> str:
        score_rows = self._get_report_account_scores(result)
        fraud_count = sum(1 for item in score_rows if int(item.get("prediction", 0)) == 1)
        normal_count = max(len(score_rows) - fraud_count, 0)
        total = max(fraud_count + normal_count, 1)
        fraud_ratio = fraud_count / total
        circumference = 2 * math.pi * 92
        fraud_dash = circumference * fraud_ratio
        normal_dash = max(circumference - fraud_dash, 0.0)
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="60" y="80" font-size="32" font-weight="700" fill="#1f4e79">预测结果分布</text>'
            '<circle cx="280" cy="270" r="92" fill="none" stroke="#e8eef6" stroke-width="48"/>'
            f'<circle cx="280" cy="270" r="92" fill="none" stroke="#e57373" stroke-width="48" stroke-linecap="round" stroke-dasharray="{fraud_dash:.2f} {circumference:.2f}" transform="rotate(-90 280 270)"/>'
            f'<circle cx="280" cy="270" r="92" fill="none" stroke="#81c784" stroke-width="48" stroke-linecap="butt" stroke-dasharray="{normal_dash:.2f} {circumference:.2f}" stroke-dashoffset="{-fraud_dash:.2f}" transform="rotate(-90 280 270)"/>'
            f'<text x="280" y="258" text-anchor="middle" font-size="38" font-weight="700" fill="#2c3e50">{len(score_rows) or 1}</text>'
            '<text x="280" y="295" text-anchor="middle" font-size="20" fill="#6b7280">账户样本</text>'
            '<rect x="510" y="170" width="26" height="26" rx="6" fill="#e57373"/>'
            f'<text x="556" y="190" font-size="24" fill="#2c3e50">欺诈账户：{fraud_count}</text>'
            '<rect x="510" y="240" width="26" height="26" rx="6" fill="#81c784"/>'
            f'<text x="556" y="260" font-size="24" fill="#2c3e50">非欺诈账户：{normal_count}</text>'
            f'<text x="510" y="338" font-size="22" fill="#6b7280">欺诈占比：{fraud_ratio * 100:.1f}%</text>'
            '</svg>'
        )
        return self._svg_to_data_uri(svg)

    def _build_confidence_distribution_svg(self, result: dict) -> str:
        score_rows = self._get_report_account_scores(result)
        buckets = {"高置信度": 0, "中置信度": 0, "低置信度": 0}
        for item in score_rows:
            label = self._stringify(item.get("confidence_label")).strip() or self._confidence_label_from_probability(
                self._safe_float(item.get("probability"))
            )
            buckets[label] = buckets.get(label, 0) + 1
        colors = {"高置信度": "#ef9a9a", "中置信度": "#ffe082", "低置信度": "#90caf9"}
        max_count = max(buckets.values(), default=1) or 1
        bars: list[str] = []
        labels: list[str] = []
        for index, key in enumerate(["高置信度", "中置信度", "低置信度"]):
            count = buckets.get(key, 0)
            height = 240 * (count / max_count)
            x = 170 + index * 190
            y = 380 - height
            bars.append(f'<rect x="{x}" y="{y:.1f}" width="90" height="{height:.1f}" rx="12" fill="{colors[key]}"/>')
            labels.append(f'<text x="{x + 45}" y="420" text-anchor="middle" font-size="24" fill="#2c3e50">{key}</text>')
            labels.append(f'<text x="{x + 45}" y="{y - 18:.1f}" text-anchor="middle" font-size="24" fill="#2c3e50">{count}</text>')
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="60" y="80" font-size="32" font-weight="700" fill="#1f4e79">样本置信度分布</text>'
            '<line x1="120" y1="380" x2="780" y2="380" stroke="#94a3b8" stroke-width="3"/>'
            '<line x1="120" y1="110" x2="120" y2="380" stroke="#94a3b8" stroke-width="3"/>'
            f'{"".join(bars)}{"".join(labels)}'
            '</svg>'
        )
        return self._svg_to_data_uri(svg)

    def _build_probability_distribution_svg(self, result: dict) -> str:
        score_rows = self._get_report_account_scores(result)
        probabilities = [min(max(self._safe_float(item.get("probability")), 0.0), 1.0) for item in score_rows]
        bins = [0, 0, 0, 0, 0]
        for probability in probabilities:
            index = min(int(probability * 5), 4)
            bins[index] += 1
        max_count = max(bins, default=1) or 1
        bars: list[str] = []
        labels: list[str] = []
        for index, count in enumerate(bins):
            height = 220 * (count / max_count)
            x = 120 + index * 125
            y = 370 - height
            bars.append(f'<rect x="{x}" y="{y:.1f}" width="78" height="{height:.1f}" rx="10" fill="#64b5f6"/>')
            labels.append(f'<text x="{x + 39}" y="410" text-anchor="middle" font-size="20" fill="#2c3e50">{index / 5:.1f}-{(index + 1) / 5:.1f}</text>')
            labels.append(f'<text x="{x + 39}" y="{y - 16:.1f}" text-anchor="middle" font-size="22" fill="#2c3e50">{count}</text>')
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="60" y="80" font-size="32" font-weight="700" fill="#1f4e79">预测概率分布图</text>'
            '<line x1="90" y1="370" x2="780" y2="370" stroke="#94a3b8" stroke-width="3"/>'
            '<line x1="90" y1="110" x2="90" y2="370" stroke="#94a3b8" stroke-width="3"/>'
            f'{"".join(bars)}{"".join(labels)}'
            '</svg>'
        )
        return self._svg_to_data_uri(svg)

    def _build_model_comparison_svg(self, result: dict) -> str:
        score_rows = self._get_report_account_scores(result)[:5]
        if not score_rows:
            score_rows = [{"account": "样本1", "gru_probability": 0.0, "xgb_probability": 0.0}]
        bars: list[str] = []
        labels: list[str] = []
        for index, item in enumerate(score_rows):
            base_x = 110 + index * 145
            gru_probability = min(max(self._safe_float(item.get("gru_probability")), 0.0), 1.0)
            xgb_probability = min(max(self._safe_float(item.get("xgb_probability")), 0.0), 1.0)
            gru_height = 230 * gru_probability
            xgb_height = 230 * xgb_probability
            bars.append(f'<rect x="{base_x}" y="{370 - gru_height:.1f}" width="34" height="{gru_height:.1f}" rx="8" fill="#64b5f6"/>')
            bars.append(f'<rect x="{base_x + 44}" y="{370 - xgb_height:.1f}" width="34" height="{xgb_height:.1f}" rx="8" fill="#ffb74d"/>')
            labels.append(f'<text x="{base_x + 39}" y="404" text-anchor="middle" font-size="18" fill="#2c3e50">{html.escape(self._trim_text(self._stringify(item.get("account")) or f"#{index + 1}", 8))}</text>')
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="60" y="80" font-size="32" font-weight="700" fill="#1f4e79">传递模块与图核模块概率对比</text>'
            '<rect x="580" y="120" width="22" height="22" rx="5" fill="#64b5f6"/><text x="614" y="138" font-size="20" fill="#2c3e50">传递模块</text>'
            '<rect x="580" y="158" width="22" height="22" rx="5" fill="#ffb74d"/><text x="614" y="176" font-size="20" fill="#2c3e50">图核模块</text>'
            '<line x1="90" y1="370" x2="810" y2="370" stroke="#94a3b8" stroke-width="3"/>'
            '<line x1="90" y1="110" x2="90" y2="370" stroke="#94a3b8" stroke-width="3"/>'
            f'{"".join(bars)}{"".join(labels)}'
            '</svg>'
        )
        return self._svg_to_data_uri(svg)

    def _build_high_risk_features_svg(self, normalized: dict, result: dict) -> str:
        score_rows = self._get_report_account_scores(result)
        target_account = score_rows[0]["account"] if score_rows else ""
        transactions = [item for item in normalized.get("standardized_transactions", []) or [] if self._stringify(item.get("zhdh")).strip() == target_account]
        if not transactions:
            transactions = list(normalized.get("standardized_transactions", []) or [])
        amounts = [self._safe_float(item.get("jyje")) for item in transactions if self._safe_float(item.get("jyje")) > 0]
        counterpart_count = len({self._stringify(item.get("dfzh")).strip() for item in transactions if self._stringify(item.get("dfzh")).strip()})
        feature_rows = [
            ("平均交易金额", sum(amounts) / len(amounts) if amounts else 0.0),
            ("最大交易金额", max(amounts, default=0.0)),
            ("关联对手数", float(counterpart_count)),
        ]
        max_value = max((item[1] for item in feature_rows), default=1.0) or 1.0
        bars: list[str] = []
        labels: list[str] = []
        for index, (label, value) in enumerate(feature_rows):
            width = 420 * (value / max_value) if max_value else 0.0
            y = 150 + index * 105
            bars.append(f'<rect x="210" y="{y}" width="{width:.1f}" height="36" rx="12" fill="#90caf9"/>')
            labels.append(f'<text x="70" y="{y + 25}" font-size="24" fill="#2c3e50">{label}</text>')
            labels.append(f'<text x="{225 + width:.1f}" y="{y + 25}" font-size="22" fill="#2c3e50">{value:.2f}</text>')
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="60" y="80" font-size="32" font-weight="700" fill="#1f4e79">高风险样本 TOP3 特征均值</text>'
            f'{"".join(bars)}{"".join(labels)}'
            '</svg>'
        )
        return self._svg_to_data_uri(svg)

    def _build_probability_violin_boxplot_svg(self, result: dict) -> str:
        score_rows = self._get_report_account_scores(result)
        probabilities = sorted(min(max(self._safe_float(item.get("probability")), 0.0), 1.0) for item in score_rows)
        if not probabilities:
            probabilities = [0.0]
        minimum = probabilities[0]
        maximum = probabilities[-1]
        median = probabilities[len(probabilities) // 2]
        q1 = probabilities[max(int((len(probabilities) - 1) * 0.25), 0)]
        q3 = probabilities[max(int((len(probabilities) - 1) * 0.75), 0)]

        def y_pos(value: float) -> float:
            return 410 - value * 250

        dots = []
        for index, value in enumerate(probabilities):
            offset = (index % 5) * 18 - 36
            dots.append(f'<circle cx="{450 + offset}" cy="{y_pos(value):.1f}" r="7" fill="#64b5f6" opacity="0.75"/>')

        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="60" y="80" font-size="32" font-weight="700" fill="#1f4e79">预测概率的箱线图与小提琴图</text>'
            '<ellipse cx="450" cy="270" rx="88" ry="140" fill="#dbeafe" opacity="0.85"/>'
            '<line x1="450" y1="160" x2="450" y2="410" stroke="#475569" stroke-width="4"/>'
            f'<line x1="450" y1="{y_pos(minimum):.1f}" x2="450" y2="{y_pos(maximum):.1f}" stroke="#334155" stroke-width="5"/>'
            f'<rect x="408" y="{y_pos(q3):.1f}" width="84" height="{max(y_pos(q1) - y_pos(q3), 12):.1f}" fill="#93c5fd" stroke="#2563eb" stroke-width="3"/>'
            f'<line x1="408" y1="{y_pos(median):.1f}" x2="492" y2="{y_pos(median):.1f}" stroke="#1e293b" stroke-width="4"/>'
            f'{"".join(dots)}'
            '<text x="560" y="200" font-size="22" fill="#2c3e50">最小值 / Q1 / 中位数 / Q3 / 最大值</text>'
            f'<text x="560" y="238" font-size="20" fill="#6b7280">{minimum:.3f} / {q1:.3f} / {median:.3f} / {q3:.3f} / {maximum:.3f}</text>'
            '</svg>'
        )
        return self._svg_to_data_uri(svg)

    def _build_correlation_heatmap_svg(self, result: dict) -> str:
        score_rows = self._get_report_account_scores(result)
        final_values = [self._safe_float(item.get("probability")) for item in score_rows]
        gru_values = [self._safe_float(item.get("gru_probability")) for item in score_rows]
        xgb_values = [self._safe_float(item.get("xgb_probability")) for item in score_rows]
        labels = ["AT-GNN", "传递模块", "图核模块"]
        matrix = [
            [1.0, self._pearson(final_values, gru_values), self._pearson(final_values, xgb_values)],
            [self._pearson(gru_values, final_values), 1.0, self._pearson(gru_values, xgb_values)],
            [self._pearson(xgb_values, final_values), self._pearson(xgb_values, gru_values), 1.0],
        ]
        blocks: list[str] = []
        texts: list[str] = []
        for row_index, row in enumerate(matrix):
            for col_index, value in enumerate(row):
                x = 240 + col_index * 130
                y = 140 + row_index * 90
                fill = self._heatmap_color(value)
                blocks.append(f'<rect x="{x}" y="{y}" width="110" height="70" rx="12" fill="{fill}"/>')
                texts.append(f'<text x="{x + 55}" y="{y + 44}" text-anchor="middle" font-size="22" font-weight="700" fill="#1f2937">{value:.2f}</text>')
        axis_labels = []
        for index, label in enumerate(labels):
            axis_labels.append(f'<text x="{295 + index * 130}" y="118" text-anchor="middle" font-size="22" fill="#2c3e50">{label}</text>')
            axis_labels.append(f'<text x="200" y="{184 + index * 90}" text-anchor="end" font-size="22" fill="#2c3e50">{label}</text>')
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="60" y="80" font-size="32" font-weight="700" fill="#1f4e79">模型概率之间的相关性热力图</text>'
            f'{"".join(blocks)}{"".join(texts)}{"".join(axis_labels)}'
            '</svg>'
        )
        return self._svg_to_data_uri(svg)

    def _build_confidence_vs_prediction_svg(self, result: dict) -> str:
        score_rows = self._get_report_account_scores(result)
        labels = ["高置信度", "中置信度", "低置信度"]
        fraud_counts = {label: 0 for label in labels}
        normal_counts = {label: 0 for label in labels}
        for item in score_rows:
            label = self._stringify(item.get("confidence_label")).strip() or self._confidence_label_from_probability(
                self._safe_float(item.get("probability"))
            )
            if label not in fraud_counts:
                continue
            if int(self._safe_float(item.get("prediction"))) == 1:
                fraud_counts[label] += 1
            else:
                normal_counts[label] += 1
        max_count = max([fraud_counts[label] + normal_counts[label] for label in labels], default=1) or 1
        bars: list[str] = []
        annotations: list[str] = []
        for index, label in enumerate(labels):
            total = fraud_counts[label] + normal_counts[label]
            base_x = 150 + index * 190
            fraud_height = 230 * (fraud_counts[label] / max_count)
            normal_height = 230 * (normal_counts[label] / max_count)
            fraud_y = 370 - fraud_height
            normal_y = fraud_y - normal_height
            bars.append(f'<rect x="{base_x}" y="{fraud_y:.1f}" width="92" height="{fraud_height:.1f}" rx="10" fill="#ef9a9a"/>')
            bars.append(f'<rect x="{base_x}" y="{normal_y:.1f}" width="92" height="{normal_height:.1f}" rx="10" fill="#a5d6a7"/>')
            annotations.append(f'<text x="{base_x + 46}" y="406" text-anchor="middle" font-size="22" fill="#2c3e50">{label}</text>')
            annotations.append(f'<text x="{base_x + 46}" y="{normal_y - 16:.1f}" text-anchor="middle" font-size="22" fill="#2c3e50">{total}</text>')
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="60" y="80" font-size="32" font-weight="700" fill="#1f4e79">置信度与预测标签的关系</text>'
            '<rect x="600" y="128" width="22" height="22" rx="5" fill="#ef9a9a"/><text x="634" y="146" font-size="20" fill="#2c3e50">欺诈账户</text>'
            '<rect x="600" y="166" width="22" height="22" rx="5" fill="#a5d6a7"/><text x="634" y="184" font-size="20" fill="#2c3e50">非欺诈账户</text>'
            '<line x1="100" y1="370" x2="800" y2="370" stroke="#94a3b8" stroke-width="3"/>'
            '<line x1="100" y1="110" x2="100" y2="370" stroke="#94a3b8" stroke-width="3"/>'
            f'{"".join(bars)}{"".join(annotations)}'
            '</svg>'
        )
        return self._svg_to_data_uri(svg)

    def _pearson(self, left: list[float], right: list[float]) -> float:
        if len(left) != len(right) or len(left) < 2:
            return 0.0
        left_mean = sum(left) / len(left)
        right_mean = sum(right) / len(right)
        numerator = sum((lx - left_mean) * (rx - right_mean) for lx, rx in zip(left, right, strict=True))
        left_denominator = sum((lx - left_mean) ** 2 for lx in left)
        right_denominator = sum((rx - right_mean) ** 2 for rx in right)
        denominator = math.sqrt(left_denominator * right_denominator)
        if denominator == 0:
            return 0.0
        return max(min(numerator / denominator, 1.0), -1.0)

    def _heatmap_color(self, value: float) -> str:
        normalized = (value + 1.0) / 2.0
        red = int(255 - (80 * normalized))
        green = int(236 - (60 * normalized))
        blue = int(246 - (180 * normalized))
        return f"rgb({red},{green},{blue})"

    def _build_signal_summary_svg(self, result: dict) -> str:
        signals = result.get("risk_signals", [])[:4] or ["暂无风险信号"]
        y = 70
        lines = []
        for signal in signals:
            lines.append(f'<text x="60" y="{y}" font-size="24" fill="#2c3e50">• {html.escape(self._stringify(signal))}</text>')
            y += 58
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<rect x="40" y="40" width="820" height="440" rx="14" fill="#eef6fc" stroke="#3498db" stroke-width="3"/>'
            '<text x="60" y="95" font-size="32" font-weight="700" fill="#1f4e79">风险信号摘要</text>'
            f'{"".join(lines)}'
            "</svg>"
        )
        return self._svg_to_data_uri(svg)

    def _build_amount_distribution_svg(self, result: dict) -> str:
        candidates = result.get("transaction_candidates", [])[:5]
        amounts = [max(self._safe_float(item.get("amount")), 0.0) for item in candidates]
        max_amount = max(amounts, default=1.0) or 1.0
        bars = []
        labels = []
        for index, amount in enumerate(amounts or [0.0], start=1):
            height = 260 * (amount / max_amount) if max_amount else 0
            x = 110 + (index - 1) * 130
            y = 390 - height
            bars.append(f'<rect x="{x}" y="{y}" width="72" height="{height}" rx="8" fill="#3498db"/>')
            labels.append(f'<text x="{x + 36}" y="430" font-size="22" text-anchor="middle" fill="#2c3e50">#{index}</text>')
            labels.append(f'<text x="{x + 36}" y="{y - 12}" font-size="20" text-anchor="middle" fill="#2c3e50">{amount:.0f}</text>')
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<line x1="80" y1="390" x2="830" y2="390" stroke="#6b7280" stroke-width="3"/>'
            '<line x1="80" y1="80" x2="80" y2="390" stroke="#6b7280" stroke-width="3"/>'
            '<text x="80" y="50" font-size="32" font-weight="700" fill="#1f4e79">候选交易金额分布</text>'
            f'{"".join(bars)}{"".join(labels)}'
            "</svg>"
        )
        return self._svg_to_data_uri(svg)

    def _build_asset_summary_svg(self, assets: list[dict]) -> str:
        counts = {
            "文档/图片": len([asset for asset in assets if asset.get("asset_type") == "document"]),
            "表格": len([asset for asset in assets if asset.get("asset_type") == "spreadsheet"]),
            "文本": len([asset for asset in assets if asset.get("asset_type") == "text"]),
        }
        colors = ["#3498db", "#2ecc71", "#f39c12"]
        y = 120
        rows = []
        for (label, count), color in zip(counts.items(), colors, strict=True):
            rows.append(f'<rect x="70" y="{y}" width="34" height="34" rx="6" fill="{color}"/>')
            rows.append(f'<text x="128" y="{y + 25}" font-size="26" fill="#2c3e50">{label}</text>')
            rows.append(f'<text x="760" y="{y + 25}" font-size="26" text-anchor="end" fill="#2c3e50">{count}</text>')
            rows.append(f'<rect x="128" y="{y + 42}" width="{count * 120 if count else 10}" height="16" rx="8" fill="{color}" opacity="0.85"/>')
            y += 110
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="70" y="68" font-size="32" font-weight="700" fill="#1f4e79">资产类型概览</text>'
            f'{"".join(rows)}'
            "</svg>"
        )
        return self._svg_to_data_uri(svg)

    def _build_account_link_svg(self, result: dict) -> str:
        nodes = result.get("link_path", [])[:4]
        if not nodes:
            nodes = [{"account": "待补充证据", "risk_level": result.get("risk_level", "low"), "action": "复核"}]
        positions = [(150, 260), (360, 160), (360, 360), (620, 260)]
        circles = []
        arrows = []
        for index, node in enumerate(nodes):
            x, y = positions[index] if index < len(positions) else (620, 260)
            color = {"high": "#e74c3c", "medium": "#f39c12", "low": "#2ecc71"}.get(
                self._normalize_risk_level_value(node.get("risk_level"), "medium"),
                "#3498db",
            )
            circles.append(f'<circle cx="{x}" cy="{y}" r="60" fill="{color}" opacity="0.88"/>')
            circles.append(f'<text x="{x}" y="{y - 5}" font-size="18" text-anchor="middle" fill="#ffffff">{html.escape(self._trim_text(self._stringify(node.get("account")), 12))}</text>')
            circles.append(f'<text x="{x}" y="{y + 22}" font-size="16" text-anchor="middle" fill="#ffffff">{html.escape(self._display_risk_level(self._normalize_risk_level_value(node.get("risk_level"), "medium")))}</text>')
            if index > 0:
                prev_x, prev_y = positions[index - 1]
                arrows.append(f'<line x1="{prev_x + 60}" y1="{prev_y}" x2="{x - 60}" y2="{y}" stroke="#7f8c8d" stroke-width="6" marker-end="url(#arrow)" />')
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<defs><marker id="arrow" markerWidth="12" markerHeight="12" refX="9" refY="6" orient="auto">'
            '<path d="M0,0 L12,6 L0,12 z" fill="#7f8c8d"/></marker></defs>'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="70" y="68" font-size="32" font-weight="700" fill="#1f4e79">重点账户链路</text>'
            f'{"".join(arrows)}{"".join(circles)}'
            "</svg>"
        )
        return self._svg_to_data_uri(svg)

    def _build_entity_summary_svg(self, normalized: dict) -> str:
        accounts = normalized.get("entities", {}).get("accounts", [])[:4]
        timestamps = normalized.get("entities", {}).get("timestamps", [])[:4]
        lines = []
        y = 120
        for title, items in (("账户", accounts), ("时间", timestamps)):
            value = "、".join(self._stringify(item) for item in items) if items else "暂无"
            lines.append(f'<text x="70" y="{y}" font-size="26" fill="#1f4e79">{title}</text>')
            lines.append(f'<text x="180" y="{y}" font-size="22" fill="#2c3e50">{html.escape(self._trim_text(value, 34))}</text>')
            y += 90
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="900" height="520" viewBox="0 0 900 520">'
            '<rect width="900" height="520" rx="16" fill="#ffffff"/>'
            '<text x="70" y="68" font-size="32" font-weight="700" fill="#1f4e79">结构化摘要</text>'
            f'{"".join(lines)}'
            "</svg>"
        )
        return self._svg_to_data_uri(svg)

    def _build_link_analysis_svg(self, result: dict) -> str:
        nodes = result.get("link_path", [])[:5]
        if not nodes:
            nodes = [{"account": "待补充证据", "risk_level": result.get("risk_level", "low"), "action": "复核"}]
        positions = [(120, 210), (300, 110), (300, 310), (520, 110), (520, 310)]
        circles = []
        arrows = []
        for index, node in enumerate(nodes):
            x, y = positions[index] if index < len(positions) else (720, 210)
            risk_level = self._normalize_risk_level_value(node.get("risk_level"), result.get("risk_level", "medium"))
            color = {"high": "#e74c3c", "medium": "#f39c12", "low": "#2ecc71"}.get(risk_level, "#3498db")
            circles.append(f'<circle cx="{x}" cy="{y}" r="56" fill="{color}" opacity="0.92"/>')
            circles.append(f'<text x="{x}" y="{y - 4}" font-size="18" text-anchor="middle" fill="#ffffff">{html.escape(self._trim_text(self._stringify(node.get("account")), 12))}</text>')
            circles.append(f'<text x="{x}" y="{y + 20}" font-size="16" text-anchor="middle" fill="#ffffff">{html.escape(self._trim_text(self._stringify(node.get("action")), 8))}</text>')
            if index > 0:
                prev_x, prev_y = positions[index - 1]
                arrows.append(f'<line x1="{prev_x + 52}" y1="{prev_y}" x2="{x - 52}" y2="{y}" stroke="#6b7280" stroke-width="5" marker-end="url(#arrow2)" />')
        svg = (
            '<svg xmlns="http://www.w3.org/2000/svg" width="1200" height="500" viewBox="0 0 1200 500">'
            '<defs><marker id="arrow2" markerWidth="12" markerHeight="12" refX="9" refY="6" orient="auto">'
            '<path d="M0,0 L12,6 L0,12 z" fill="#6b7280"/></marker></defs>'
            '<rect width="1200" height="500" rx="18" fill="#fdfefe"/>'
            '<rect x="20" y="20" width="1160" height="460" rx="18" fill="#ffffff" stroke="#d8dee9" stroke-width="3"/>'
            f'{"".join(arrows)}{"".join(circles)}'
            "</svg>"
        )
        return self._svg_to_data_uri(svg)

    def _write_pdf_report(self, report_path: Path, pdf_path: Path) -> None:
        edge_executable = self._find_edge_executable()
        if edge_executable is None:
            raise RuntimeError("未找到 Microsoft Edge，无法导出 PDF 报告")

        profile_dir = report_path.parent / ".edge-pdf-profile"
        profile_dir.mkdir(parents=True, exist_ok=True)
        command = [
            str(edge_executable),
            "--headless",
            "--disable-gpu",
            "--allow-file-access-from-files",
            "--no-first-run",
            "--no-default-browser-check",
            f"--user-data-dir={profile_dir}",
            "--no-pdf-header-footer",
            f"--print-to-pdf={pdf_path.resolve()}",
            report_path.resolve().as_uri(),
        ]
        completed = subprocess.run(
            command,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="ignore",
            timeout=120,
            check=False,
        )
        if completed.returncode != 0 or not pdf_path.exists():
            detail = (completed.stderr or completed.stdout or "").strip()
            raise RuntimeError(detail or "PDF 报告生成失败")

    def _find_edge_executable(self) -> Path | None:
        candidates = [
            Path(r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"),
            Path(r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"),
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return None

    def _file_to_data_uri(self, path: Path, mime_type: str | None = None) -> str:
        resolved_mime = mime_type or self._guess_mime_type(path)
        encoded = base64.b64encode(path.read_bytes()).decode("ascii")
        return f"data:{resolved_mime};base64,{encoded}"

    def _svg_to_data_uri(self, svg: str) -> str:
        encoded = base64.b64encode(svg.encode("utf-8")).decode("ascii")
        return f"data:image/svg+xml;base64,{encoded}"

    def _trim_text(self, value: str, max_length: int) -> str:
        if len(value) <= max_length:
            return value
        return value[: max_length - 1] + "…"

    def _to_json_safe(self, value: object) -> object:
        if isinstance(value, dict):
            return {self._stringify(key): self._to_json_safe(item) for key, item in value.items()}
        if isinstance(value, (list, tuple, set)):
            return [self._to_json_safe(item) for item in value]
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, dt_time):
            return value.strftime("%H:%M:%S")
        return value

    def _select_columns_for_normalization(self, columns: list[str]) -> list[str]:
        selected: list[str] = []
        keyword_groups = (
            ("付款", "payer", "转出", "付款方", "付款账号", "付款账户", "账户代号", "账户", "账号"),
            ("收款", "receiver", "转入", "收款方", "对方账户", "对方账号", "对手账户", "对方"),
            ("金额", "amount", "交易额"),
            ("日期", "date"),
            ("时间", "time", "timestamp"),
            ("余额", "balance"),
            ("渠道", "channel"),
            ("银行", "行号", "bank"),
        )
        for keywords in keyword_groups:
            matched = self._find_matching_key(columns, keywords)
            if matched and matched not in selected:
                selected.append(matched)
        return selected or columns[:8]

    def _build_normalize_payload(self, bundle: dict) -> dict:
        payload = {
            "job_id": bundle.get("job_id"),
            "documents_markdown": [],
            "structured_tables": [],
            "plain_texts": [],
            "source_meta": self._to_json_safe(bundle.get("source_meta", [])),
        }

        for item in bundle.get("documents_markdown", []):
            payload["documents_markdown"].append(
                {
                    "asset_id": item.get("asset_id"),
                    "source_type": item.get("source_type"),
                    "markdown": self._trim_text(self._stringify(item.get("markdown")), 6000),
                }
            )

        for item in bundle.get("plain_texts", []):
            payload["plain_texts"].append(
                {
                    "asset_id": item.get("asset_id"),
                    "content": self._trim_text(self._stringify(item.get("content")), 6000),
                }
            )

        for table in bundle.get("structured_tables", []):
            columns = [self._stringify(column) for column in table.get("columns", [])]
            selected_columns = self._select_columns_for_normalization(columns)
            compact_rows = []
            for row in table.get("rows", []):
                compact_rows.append([self._to_json_safe(row.get(column)) for column in selected_columns])
            payload["structured_tables"].append(
                {
                    "asset_id": table.get("asset_id"),
                    "sheet_name": table.get("sheet_name"),
                    "row_count": len(table.get("rows", [])),
                    "selected_columns": selected_columns,
                    "rows": compact_rows,
                }
            )

        return payload

    def _build_normalize_prompt(self, bundle: dict) -> list[dict]:
        normalize_payload = self._build_normalize_payload(bundle)
        example_input = {
            "job_id": "example-voucher",
            "documents_markdown": [
                {
                    "asset_id": "voucher-1",
                    "source_type": "document",
                    "markdown": (
                        "# 转账凭证单\n\n"
                        "交易日期：2020/3/1 交易时间：4:11\n\n"
                        "交易流水序号：123\n\n"
                        "<table>"
                        "<tr><td>账户代号</td><td colspan=\"2\">324255555555565</td></tr>"
                        "<tr><td>对方账户</td><td colspan=\"2\">8848884888488840</td></tr>"
                        "<tr><td>交易金额</td><td>2554.45</td><td>账户余额：356.1</td></tr>"
                        "<tr><td>借贷标记</td><td>是</td></tr>"
                        "<tr><td>对方行号</td><td>c01</td></tr>"
                        "<tr><td>交易渠道</td><td>网银转账</td></tr>"
                        "<tr><td>年龄</td><td>22</td></tr>"
                        "<tr><td>性别</td><td>男</td></tr>"
                        "</table>"
                    ),
                }
            ],
            "structured_tables": [],
            "plain_texts": [],
            "source_meta": [{"asset_id": "voucher-1", "asset_type": "document", "original_name": "转账凭证.png"}],
        }
        example_output = {
            "standardized_transactions": [
                {
                    "jylsxh": "123",
                    "zhdh": "324255555555565",
                    "dfzh": "8848884888488840",
                    "jdbj": 1,
                    "jyje": 2554.45,
                    "zhye": 356.1,
                    "dfhh": "c01",
                    "jyrq": "2020-03-01",
                    "jysj": "04:11:00",
                    "jyqd": "网银转账",
                    "dfmccd": 0,
                    "xb": 1,
                    "年龄": 22
                }
            ],
            "transaction_candidates": [
                {
                    "payer_account": "324255555555565",
                    "receiver_account": "8848884888488840",
                    "amount": 2554.45,
                    "timestamp": "2020-03-01 04:11:00",
                    "channel": "网银转账",
                    "description": "转账凭证提取结果",
                    "evidence_refs": ["voucher-1"]
                }
            ],
            "entities": {
                "accounts": ["324255555555565", "8848884888488840"],
                "names": [],
                "amounts": [2554.45, 356.1],
                "timestamps": ["2020-03-01 04:11:00"],
                "locations": []
            },
            "risk_signals": ["存在文档类多模态证据"],
            "evidence_alignment": [
                {
                    "asset_id": "voucher-1",
                    "mapped_fields": ["jylsxh", "zhdh", "dfzh", "jdbj", "jyje", "zhye", "dfhh", "jyrq", "jysj", "jyqd", "xb", "年龄"]
                }
            ],
            "summary": "已从转账凭证中提取出一条标准交易记录。"
        }
        return [
            {
                "role": "system",
                "content": (
                    "你是金融多模态证据归一化引擎。你的任务不是做最终欺诈判断，而是从干净的 Markdown、文本中抽取标准交易记录、实体、风险信号。"
                    "只输出合法 JSON，不要输出解释。"
                    "严格要求："
                    "1. 如果凭证里出现了账户代号、对方账户、交易金额、日期/时间、借贷标记等字段，必须优先生成 standardized_transactions。"
                    "2. entities.amounts 只允许保留真实金额字段，例如交易金额、账户余额；绝不能放日期拆分数字、流水号、账号、页码、坐标、bbox。"
                    "3. accounts 只允许保留账号；timestamps 只允许保留时间戳；不要把任何无关数字混进去。"
                    "4. 缺失字段可以留空或填 0，但不要捏造。"
                ),
            },
            {
                "role": "user",
                "content": (
                    "下面是一个示例。\n\n"
                    f"输入：{json.dumps(example_input, ensure_ascii=False)}\n\n"
                    f"输出：{json.dumps(example_output, ensure_ascii=False)}"
                ),
            },
            {
                "role": "user",
                "content": (
                    "请将以下证据包整理成 JSON，字段必须包含：standardized_transactions, transaction_candidates, entities, risk_signals, evidence_alignment, summary。\n\n"
                    f"{json.dumps(normalize_payload, ensure_ascii=False)}"
                ),
            },
        ]

    def _build_analysis_prompt(self, normalized: dict) -> list[dict]:
        json_safe_normalized = self._to_json_safe(normalized)
        return [
            {
                "role": "system",
                "content": (
                    "你是金融欺诈分析助手。请基于结构化证据输出最终风险结论。"
                    "只输出合法 JSON，字段必须包含：risk_level, confidence, model_source, narrative, suggested_actions, link_path。"
                    "link_path 为数组，元素必须包含 account, risk_level, action。"
                ),
            },
            {"role": "user", "content": json.dumps(json_safe_normalized, ensure_ascii=False)},
        ]

    def _call_chat_completion(self, messages: list[dict]) -> str:
        response = self._request_json(
            method="POST",
            url=f"{settings.analysis_llm_base_url.rstrip('/')}/chat/completions",
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {settings.analysis_llm_api_key}",
            },
            data={
                "model": settings.analysis_llm_model,
                "messages": messages,
                "temperature": 0.1,
                "response_format": {"type": "json_object"},
            },
            timeout=settings.analysis_llm_timeout_seconds,
        )
        choice = (response.get("choices") or [{}])[0]
        message = choice.get("message", {})
        content = message.get("content", "")
        if isinstance(content, list):
            return "".join(part.get("text", "") for part in content if isinstance(part, dict))
        return content

    def _extract_json_object(self, payload: str) -> dict:
        payload = payload.strip()
        if not payload:
            raise ValueError("模型返回为空")
        try:
            return json.loads(payload)
        except json.JSONDecodeError:
            start = payload.find("{")
            end = payload.rfind("}")
            if start == -1 or end == -1 or end <= start:
                raise
            return json.loads(payload[start : end + 1])

    def _request_json(self, *, method: str, url: str, headers: dict, data: dict | None = None, timeout: int | float) -> dict:
        response = self._request_with_retry(
            method=method,
            url=url,
            headers=headers,
            json=data,
            timeout=timeout,
        )
        try:
            return response.json()
        except ValueError as exc:
            raise RuntimeError(f"外部服务返回的不是合法 JSON：{response.text[:200]}") from exc

    def _put_file(self, upload_url: str, path: Path) -> None:
        self._request_with_retry(
            method="PUT",
            url=upload_url,
            data=path.read_bytes(),
            timeout=settings.analysis_llm_timeout_seconds,
        )

    def _download_bytes(self, url: str, timeout: int | float) -> bytes:
        response = self._request_with_retry(method="GET", url=url, timeout=timeout)
        return response.content

    def _request_with_retry(self, *, method: str, url: str, timeout: int | float, **kwargs) -> requests.Response:
        last_error: Exception | None = None
        for index, delay in enumerate(HTTP_RETRY_DELAYS, start=1):
            if delay > 0:
                time.sleep(delay)
            try:
                response = self.http.request(method=method, url=url, timeout=timeout, **kwargs)
                response.raise_for_status()
                return response
            except requests.HTTPError as exc:
                status_code = exc.response.status_code if exc.response is not None else None
                if status_code is not None and 400 <= status_code < 500:
                    detail = exc.response.text if exc.response is not None else ""
                    raise RuntimeError(detail or str(exc)) from exc
                last_error = exc
            except requests.RequestException as exc:
                last_error = exc

        raise RuntimeError(str(last_error) if last_error is not None else "外部请求失败")

    def _write_temp_zip(self, asset_id: str, content: bytes) -> Path:
        temp_dir = self.storage_dir / "_mineru_cache"
        temp_dir.mkdir(parents=True, exist_ok=True)
        zip_path = temp_dir / f"{asset_id}.zip"
        zip_path.write_bytes(content)
        return zip_path

    def _find_matching_key(self, columns: list[str], keywords: tuple[str, ...]) -> str | None:
        lowered = [(column or "").lower() for column in columns]
        for keyword in keywords:
            keyword_lower = keyword.lower()
            for index, column in enumerate(lowered):
                if keyword_lower in column:
                    return columns[index]
        return None

    def _table_to_markdown(self, sheet_name: str, headers: list[str], rows: list[tuple | list]) -> str:
        safe_headers = [header or "-" for header in headers]
        if not safe_headers:
            return f"## {sheet_name}\n\n空表"
        lines = [f"## {sheet_name}", "", "| " + " | ".join(safe_headers) + " |"]
        lines.append("| " + " | ".join("---" for _ in safe_headers) + " |")
        for row in rows:
            values = [self._stringify(row[index]) if index < len(row) else "" for index in range(len(safe_headers))]
            lines.append("| " + " | ".join(values) + " |")
        return "\n".join(lines)

    def _classify_asset_type(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix in SPREADSHEET_EXTENSIONS:
            return "spreadsheet"
        if suffix in TEXT_EXTENSIONS:
            return "text"
        return "document"

    def _guess_mime_type(self, path: Path) -> str:
        return mimetypes.guess_type(str(path))[0] or "application/octet-stream"

    def _generate_job_id(self) -> str:
        return f"analysis-{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6]}"

    def _generate_asset_id(self) -> str:
        return f"asset-{uuid.uuid4().hex}"

    def _to_asset_schema(self, item: dict) -> AnalysisAsset:
        return AnalysisAsset(
            asset_id=item["asset_id"],
            asset_type=item["asset_type"],
            original_name=item["original_name"],
            mime_type=item["mime_type"],
            size_bytes=item["size_bytes"],
        )

    def _ensure_job_access(self, job: dict, current_user: dict) -> None:
        if current_user["role"] != "admin" and job["created_by"] != current_user["username"]:
            raise AppError("当前账户无权访问该分析任务", status_code=403, code="FORBIDDEN")

    def _now_iso(self) -> str:
        return datetime.utcnow().replace(microsecond=0).isoformat()

    def _safe_float(self, value: object) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    def _stringify(self, value: object) -> str:
        if value is None:
            return ""
        return str(value)

    def _normalize_direction_value(self, value: object) -> int:
        normalized = self._stringify(value).strip().lower()
        if normalized in {"1", "true", "yes", "y", "in", "credit", "贷", "入", "收入", "是"}:
            return 1
        if normalized in {"0", "false", "no", "n", "out", "debit", "借", "出", "支出", "否"}:
            return 0
        return int(self._safe_float(value)) if self._stringify(value).strip() else 0

    def _normalize_gender_value(self, value: object) -> int:
        normalized = self._stringify(value).strip().lower()
        if normalized in {"1", "male", "m", "男"}:
            return 1
        if normalized in {"0", "female", "f", "女"}:
            return 0
        return int(self._safe_float(value)) if self._stringify(value).strip() else 0

    def _normalize_date_value(self, value: object) -> str:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        text = self._stringify(value).strip()
        if not text:
            return ""
        if len(text) >= 10:
            return text[:10]
        return text

    def _normalize_time_value(self, value: object) -> str:
        if isinstance(value, datetime):
            return value.strftime("%H:%M:%S")
        if isinstance(value, dt_time):
            return value.strftime("%H:%M:%S")
        text = self._stringify(value).strip()
        if not text:
            return ""
        if len(text) >= 8:
            return text[-8:]
        return text
    def _normalize_result_payload(self, payload: dict | None) -> dict | None:
        if not isinstance(payload, dict):
            return None

        normalized = dict(payload)
        fallback_level = self._normalize_risk_level_value(normalized.get("risk_level"), "medium")
        normalized["risk_level"] = fallback_level
        normalized["confidence"] = self._safe_float(normalized.get("confidence")) or 0.0
        normalized["model_source"] = self._stringify(normalized.get("model_source")) or "heuristic-fallback"
        normalized["narrative"] = self._stringify(normalized.get("narrative")) or "系统已完成初步判定。"
        normalized["normalized_summary"] = self._stringify(normalized.get("normalized_summary"))
        normalized["transaction_candidates"] = (
            normalized.get("transaction_candidates")
            if isinstance(normalized.get("transaction_candidates"), list)
            else []
        )

        normalized_actions: list[str] = []
        for item in normalized.get("suggested_actions") or []:
            if isinstance(item, str):
                action_text = item.strip()
            elif isinstance(item, dict):
                action_text = self._stringify(item.get("action") or item.get("label") or item.get("name")).strip()
            else:
                action_text = self._stringify(item).strip()
            if action_text:
                normalized_actions.append(action_text)
        normalized["suggested_actions"] = normalized_actions or self._build_actions(fallback_level)

        normalized_signals: list[str] = []
        for item in normalized.get("risk_signals") or []:
            signal_text = self._stringify(item).strip()
            if signal_text:
                normalized_signals.append(signal_text)
        normalized["risk_signals"] = normalized_signals

        normalized_path: list[dict] = []
        for item in normalized.get("link_path") or []:
            if not isinstance(item, dict):
                continue
            account = self._stringify(item.get("account")).strip() or "待确认账户"
            action = self._stringify(item.get("action")).strip() or "复核"
            risk_level = self._normalize_risk_level_value(item.get("risk_level"), fallback_level)
            normalized_path.append(
                AnalysisRiskNode(
                    account=account,
                    risk_level=risk_level,
                    action=action,
                ).model_dump()
            )
        if not normalized_path:
            normalized_path = [
                AnalysisRiskNode(
                    account="待补充证据",
                    risk_level=fallback_level,
                    action="复核",
                ).model_dump()
            ]
        normalized["link_path"] = normalized_path

        return normalized

    def _apply_focus_targets(self, normalized: dict, result: dict) -> dict:
        if not isinstance(result, dict):
            return result

        focus_accounts = {
            self._stringify(item.get("account")).strip()
            for item in focus_repository.list_targets()
            if self._stringify(item.get("account")).strip()
        }
        if not focus_accounts:
            return result

        matched_accounts: list[str] = []
        seen_accounts: set[str] = set()

        def collect(account: object) -> None:
            text = self._stringify(account).strip()
            if text and text in focus_accounts and text not in seen_accounts:
                seen_accounts.add(text)
                matched_accounts.append(text)

        for account in (normalized.get("entities") or {}).get("accounts", []):
            collect(account)

        for row in normalized.get("standardized_transactions") or []:
            collect(row.get("zhdh"))
            collect(row.get("dfzh"))

        for row in normalized.get("transaction_candidates") or []:
            collect(row.get("payer_account"))
            collect(row.get("receiver_account"))

        for row in result.get("link_path") or []:
            if isinstance(row, dict):
                collect(row.get("account"))

        if not matched_accounts:
            return result

        enriched = dict(result)
        risk_signals = [
            self._stringify(item).strip()
            for item in enriched.get("risk_signals", [])
            if self._stringify(item).strip()
        ]
        focus_signal = f"命中重点关注账号：{'、'.join(matched_accounts)}"
        if focus_signal not in risk_signals:
            risk_signals.insert(0, focus_signal)
        enriched["risk_signals"] = risk_signals

        suggested_actions = [
            self._stringify(item).strip()
            for item in enriched.get("suggested_actions", [])
            if self._stringify(item).strip()
        ]
        focus_action = "优先复核重点关注账号及其关联交易链路"
        if focus_action not in suggested_actions:
            suggested_actions.insert(0, focus_action)
        enriched["suggested_actions"] = suggested_actions

        narrative = self._stringify(enriched.get("narrative")).strip()
        focus_prefix = f"本次分析命中重点关注账号：{'、'.join(matched_accounts)}。"
        if focus_prefix not in narrative:
            enriched["narrative"] = f"{focus_prefix}{narrative}" if narrative else focus_prefix

        return enriched

    def _normalize_risk_level_value(self, value: object, fallback: str) -> str:
        normalized = self._stringify(value).strip().lower()
        mapping = {
            "high": "high",
            "medium": "medium",
            "low": "low",
            "critical": "high",
            "高": "high",
            "高风险": "high",
            "中": "medium",
            "中风险": "medium",
            "低": "low",
            "低风险": "low",
        }
        return mapping.get(normalized, fallback)

    def _display_risk_level(self, value: str) -> str:
        return {
            "high": "高风险",
            "medium": "中风险",
            "low": "低风险",
        }.get(value, value)

    def _build_actions(self, risk_level: str) -> list[str]:
        if risk_level == "high":
            return ["冻结高风险账户", "拉取关联交易链路", "生成案件报告并提交人工复核"]
        if risk_level == "medium":
            return ["继续重点监测", "补充核验交易凭证", "安排分析员复核"]
        return ["记录本次分析结果", "保留后续复查入口"]


service = AnalysisService()

